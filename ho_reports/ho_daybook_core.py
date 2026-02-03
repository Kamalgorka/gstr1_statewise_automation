import os
import re
import zipfile
import tempfile
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

# =========================
# SETTINGS (same as your script)
# =========================
TRANSACTION_WISE_STATEMENTS = {"ICICI 7021"}

BANK_NAME_MAP = {
    "ICICI 7021": "ICICI Bank - 008205007021 - C",
    "ICICI 6086": "ICICI Bank - 008205006086 - C",
    # ✅ your COA exact bank ledger name
    "AXIS": "Axis Bank- 91402006845861 - C",
}

DEFAULT_DEPARTMENT = "HO001"
DEFAULT_BRANCH_ID = "HO001"

BANK_LEDGER_ACCOUNTS = {
    "311701", "311702", "311703", "311704", "311705", "311706", "311707", "311708",
    "311709", "311710", "311711", "311712", "311713", "311714", "311715", "311716",
    "311717", "311718", "311719", "311720",
}

# =========================
# HELPERS (same as your script)
# =========================
def find_header_row_xlsx(file_path, sheet_name=None, required_headers=None, max_scan_rows=80):
    preview = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=max_scan_rows)
    required_headers = [h.strip().lower() for h in (required_headers or [])]
    for i in range(len(preview)):
        row_vals = preview.iloc[i].astype(str).str.strip().str.lower().tolist()
        if all(any(req == cell for cell in row_vals) for req in required_headers):
            return i
    return None

def clean_text(x):
    if pd.isna(x):
        return ""
    return str(x).strip()

def safe_sheet_name(name: str, max_len: int = 31) -> str:
    bad = [":", "\\", "/", "?", "*", "[", "]"]
    for b in bad:
        name = name.replace(b, " ")
    name = " ".join(name.split())
    if len(name) > max_len:
        name = name[:max_len]
    return name

def excel_hyperlink_sheet(sheet_name: str) -> str:
    return sheet_name.replace("'", "''")

def build_keyword_table(coa_df: pd.DataFrame) -> pd.DataFrame:
    coa_df.columns = [c.strip() for c in coa_df.columns]
    if "Account Name" not in coa_df.columns:
        raise ValueError("COA.xlsx must have column: 'Account Name'")
    refer_cols = [c for c in coa_df.columns if c.upper().startswith("REFER")]
    if not refer_cols:
        raise ValueError("No Refer columns found in COA (Refer1..ReferN)")
    keyword_rows = []
    for _, row in coa_df.iterrows():
        ledger = clean_text(row.get("Account Name"))
        if not ledger:
            continue
        for ref_col in refer_cols:
            kw = clean_text(row.get(ref_col))
            if kw:
                keyword_rows.append({"ledger": ledger, "keyword": kw})
    kw_df = pd.DataFrame(keyword_rows)
    if kw_df.empty:
        raise ValueError("No keywords found in Refer columns.")
    kw_df["kw_len"] = kw_df["keyword"].str.len()
    kw_df = kw_df.sort_values("kw_len", ascending=False).reset_index(drop=True)
    return kw_df

def map_ledger_from_keywords(description: str, kw_df: pd.DataFrame) -> str:
    d = str(description).upper()
    for _, r in kw_df.iterrows():
        if r["keyword"].upper() in d:
            return r["ledger"]
    return "UNMAPPED"

def move_others_last_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.copy()
    df["_is_others"] = df["Ledger"].astype(str).str.strip().str.upper().eq("OTHERS")
    df = df.sort_values(["_is_others"], ascending=[True]).drop(columns=["_is_others"])
    return df.reset_index(drop=True)

def make_daybook_layout_from_lists(receipt_list_df: pd.DataFrame, payment_list_df: pd.DataFrame) -> pd.DataFrame:
    receipt_list_df = receipt_list_df.reset_index(drop=True)
    payment_list_df = payment_list_df.reset_index(drop=True)
    max_len = max(len(receipt_list_df), len(payment_list_df))
    rows = []
    for i in range(max_len):
        r_led = receipt_list_df.loc[i, "Ledger"] if i < len(receipt_list_df) else ""
        r_amt = receipt_list_df.loc[i, "Amount"] if i < len(receipt_list_df) else ""
        p_led = payment_list_df.loc[i, "Ledger"] if i < len(payment_list_df) else ""
        p_amt = payment_list_df.loc[i, "Amount"] if i < len(payment_list_df) else ""
        rows.append([r_led, r_amt, p_led, p_amt])
    return pd.DataFrame(rows, columns=["RECEIPT", "AMOUNT", "PAYMENTS", "AMOUNT"])

def build_coa_lookup(coa_df: pd.DataFrame):
    coa_df = coa_df.copy()
    coa_df.columns = [c.strip() for c in coa_df.columns]
    for needed in ["Account Name", "Account", "Sub Account"]:
        if needed not in coa_df.columns:
            raise ValueError(f"COA.xlsx must have column: '{needed}'")
    lookup_by_name = {}
    lookup_name_by_pair = {}
    for _, r in coa_df.iterrows():
        name = clean_text(r["Account Name"])
        acc = clean_text(r["Account"])
        sub = clean_text(r["Sub Account"])
        if name:
            lookup_by_name[name] = (acc, sub)
        if acc != "" or sub != "":
            lookup_name_by_pair[(acc, sub)] = name
    return lookup_by_name, lookup_name_by_pair

def pick_document_date(bank_df: pd.DataFrame):
    for col in ["Value Date", "Txn Posted Date", "Document Date", "Tran Date"]:
        if col in bank_df.columns:
            ser = pd.to_datetime(bank_df[col], errors="coerce").dropna()
            if len(ser):
                return ser.iloc[0].date()
    return None

def is_bank_ledger_pair(acc: str, sub: str) -> bool:
    acc = clean_text(acc)
    return acc in BANK_LEDGER_ACCOUNTS

def bank_short_name_from_account_name(account_name: str) -> str:
    s = clean_text(account_name)
    s_up = s.upper()
    digit_runs = re.findall(r"\d{4,}", s)
    last4 = digit_runs[-1][-4:] if digit_runs else ""
    brand = None
    for b in ["ICICI", "HDFC", "AXIS", "SBI", "IDBI", "PNB", "CANARA", "FEDERAL", "BANDHAN", "YES"]:
        if b in s_up:
            brand = b
            break
    if brand and last4:
        return f"{brand.title()} {last4}".replace("Icici", "ICICI").replace("Sbi", "SBI").replace("Pnb", "PNB")
    if last4:
        return last4
    return s

def make_cv_reference(to_bank_name: str, from_bank_name: str) -> str:
    to_short = bank_short_name_from_account_name(to_bank_name)
    from_short = bank_short_name_from_account_name(from_bank_name)
    return f"BEING ONLINE FUND TRANSFERRED TO {to_short} FROM {from_short}"

def create_entry_df(daybook_df: pd.DataFrame, bank_display_name: str, coa_lookup_by_name: dict, coa_name_by_pair: dict, doc_date):
    bank_acc, bank_sub = coa_lookup_by_name.get(bank_display_name, ("", ""))
    if not bank_acc:
        raise ValueError(f"Bank ledger not found in COA for: {bank_display_name}")

    doc_date_val = doc_date.strftime("%d-%b-%y") if doc_date else ""
    rows = []

    def add_pair(journal_code, ref_text, debit_acc, debit_sub, credit_acc, credit_sub, amt):
        rows.append({
            "Journal Code": journal_code, "Sequence": 1,
            "Account": debit_acc, "Sub Account": debit_sub,
            "Department": DEFAULT_DEPARTMENT, "Document Date": doc_date_val,
            "Debit": amt, "Credit": "",
            "Supplier Id": "", "Customer Id": "", "SAC/HSN": "",
            "Reference": ref_text, "Branch Id": DEFAULT_BRANCH_ID,
            "Invoice Num": "", "Comments": ""
        })
        rows.append({
            "Journal Code": journal_code, "Sequence": 2,
            "Account": credit_acc, "Sub Account": credit_sub,
            "Department": DEFAULT_DEPARTMENT, "Document Date": doc_date_val,
            "Debit": "", "Credit": amt,
            "Supplier Id": "", "Customer Id": "", "SAC/HSN": "",
            "Reference": ref_text, "Branch Id": DEFAULT_BRANCH_ID,
            "Invoice Num": "", "Comments": ""
        })

    df = daybook_df.copy()
    if df.columns.tolist() == ["RECEIPT", "AMOUNT", "PAYMENTS", "AMOUNT"]:
        df.columns = ["RECEIPT", "AMOUNT", "PAYMENTS", "AMOUNT.1"]

    # RECEIPT
    for _, r in df.iterrows():
        ledger_name = clean_text(r.get("RECEIPT"))
        if not ledger_name or ledger_name.upper() in ["TOTAL"]:
            continue
        try:
            amt_val = float(r.get("AMOUNT"))
        except Exception:
            continue
        if abs(amt_val) < 0.01:
            continue

        led_acc, led_sub = coa_lookup_by_name.get(ledger_name, ("", ""))
        is_cv = is_bank_ledger_pair(bank_acc, bank_sub) and is_bank_ledger_pair(led_acc, led_sub)

        if is_cv:
            other_bank_name = coa_name_by_pair.get((led_acc, led_sub), ledger_name)
            ref_text = make_cv_reference(to_bank_name=bank_display_name, from_bank_name=other_bank_name)
            journal = "CV"
        else:
            ref_text = f"BEING {ledger_name} RECEIPT IN {bank_display_name}"
            journal = "BR"

        add_pair(journal, ref_text, bank_acc, bank_sub, led_acc, led_sub, amt_val)

    # PAYMENT
    for _, r in df.iterrows():
        ledger_name = clean_text(r.get("PAYMENTS"))
        if not ledger_name or ledger_name.upper() in ["TOTAL"]:
            continue
        try:
            amt_val = float(r.get("AMOUNT.1"))
        except Exception:
            continue
        if abs(amt_val) < 0.01:
            continue

        led_acc, led_sub = coa_lookup_by_name.get(ledger_name, ("", ""))
        is_cv = is_bank_ledger_pair(led_acc, led_sub) and is_bank_ledger_pair(bank_acc, bank_sub)

        if is_cv:
            other_bank_name = coa_name_by_pair.get((led_acc, led_sub), ledger_name)
            ref_text = make_cv_reference(to_bank_name=other_bank_name, from_bank_name=bank_display_name)
            journal = "CV"
        else:
            ref_text = f"BEING {ledger_name} PAYMENT FROM {bank_display_name}"
            journal = "BP"

        add_pair(journal, ref_text, led_acc, led_sub, bank_acc, bank_sub, amt_val)

    entry_cols = [
        "Journal Code", "Sequence", "Account", "Sub Account", "Department", "Document Date",
        "Debit", "Credit", "Supplier Id", "Customer Id", "SAC/HSN", "Reference",
        "Branch Id", "Invoice Num", "Comments"
    ]
    return pd.DataFrame(rows, columns=entry_cols)

def detect_statement_columns(bank: pd.DataFrame, base_name: str):
    cols = [str(c).strip() for c in bank.columns]
    bank.columns = cols

    # ICICI
    if "Description" in cols and "Cr/Dr" in cols:
        desc_col = "Description"
        crdr_col = "Cr/Dr"
        amt_candidates = [c for c in cols if "Transaction Amount" in c]
        if not amt_candidates:
            raise ValueError(f"{base_name}: Could not find 'Transaction Amount(INR)' column.")
        amt_col = amt_candidates[0]
        return desc_col, crdr_col, amt_col

    # AXIS
    if "Transaction Particulars" in cols:
        desc_col = "Transaction Particulars"

        crdr_candidates = [c for c in cols if c.replace(" ", "").upper() in ["DR|CR", "DR/CR", "DRCR", "DR|CR."]]
        if not crdr_candidates:
            crdr_candidates = [c for c in cols if ("DR" in c.upper() and "CR" in c.upper())]
        if not crdr_candidates:
            raise ValueError(f"{base_name}: Could not find 'DR|CR' column in AXIS statement.")
        crdr_col = crdr_candidates[0]

        amt_candidates = [c for c in cols if "AMOUNT" in c.upper() and "INR" in c.upper() and "BALANCE" not in c.upper()]
        if not amt_candidates:
            raise ValueError(f"{base_name}: Could not find 'Amount(INR)' column in AXIS statement.")
        amt_col = amt_candidates[0]

        return desc_col, crdr_col, amt_col

    raise ValueError(f"{base_name}: Unknown statement format.")

def add_all_sheets_for_statement(wb: Workbook, statement_path: str, kw_df: pd.DataFrame, coa_lookup_by_name: dict, coa_name_by_pair: dict):
    base = os.path.splitext(os.path.basename(statement_path))[0]
    display_name = BANK_NAME_MAP.get(base, base)
    title_text = display_name
    is_transaction_wise = base in TRANSACTION_WISE_STATEMENTS

    header_row = find_header_row_xlsx(statement_path, sheet_name=0, required_headers=["Description", "Cr/Dr"])
    if header_row is None:
        header_row = find_header_row_xlsx(statement_path, sheet_name=0, required_headers=["Transaction Particulars", "DR|CR"])
    if header_row is None:
        header_row = 6

    bank = pd.read_excel(statement_path, sheet_name=0, header=header_row)
    bank.columns = [str(c).strip() for c in bank.columns]

    desc_col, crdr_col, amt_col = detect_statement_columns(bank, base)

    bank[desc_col] = bank[desc_col].astype(str)
    bank[crdr_col] = bank[crdr_col].astype(str).str.strip().str.upper()
    bank[amt_col] = pd.to_numeric(bank[amt_col], errors="coerce").fillna(0)

    bank["Ledger"] = bank[desc_col].apply(lambda x: map_ledger_from_keywords(x, kw_df))
    total_cr = bank.loc[bank[crdr_col] == "CR", amt_col].sum()
    total_dr = bank.loc[bank[crdr_col] == "DR", amt_col].sum()

    # DayBook build (same)
    if not is_transaction_wise:
        bank_cr = bank[bank[crdr_col] == "CR"].copy()
        receipts = bank_cr.groupby("Ledger", as_index=False)[amt_col].sum()
        receipts = receipts[receipts["Ledger"] != "UNMAPPED"].rename(columns={amt_col: "Amount"})
        mapped_receipt_total = receipts["Amount"].sum() if len(receipts) else 0.0
        receipt_diff = round(total_cr - mapped_receipt_total, 2)
        if abs(receipt_diff) > 0.01:
            receipts = pd.concat([receipts, pd.DataFrame([{"Ledger": "Others", "Amount": receipt_diff}])], ignore_index=True)

        bank_dr = bank[bank[crdr_col] == "DR"].copy()
        payments = bank_dr.groupby("Ledger", as_index=False)[amt_col].sum()
        payments = payments[payments["Ledger"] != "UNMAPPED"].rename(columns={amt_col: "Amount"})
        mapped_payment_total = payments["Amount"].sum() if len(payments) else 0.0
        payment_diff = round(total_dr - mapped_payment_total, 2)
        if abs(payment_diff) > 0.01:
            payments = pd.concat([payments, pd.DataFrame([{"Ledger": "Others", "Amount": payment_diff}])], ignore_index=True)

        receipts = move_others_last_df(receipts.sort_values("Ledger").reset_index(drop=True))
        payments = move_others_last_df(payments.sort_values("Ledger").reset_index(drop=True))
        daybook_df = make_daybook_layout_from_lists(receipts, payments)
    else:
        cr_list = bank[(bank[crdr_col] == "CR") & (bank["Ledger"].str.upper() != "UNMAPPED")][["Ledger", amt_col]].copy()
        cr_list = cr_list.rename(columns={amt_col: "Amount"})
        dr_list = bank[(bank[crdr_col] == "DR") & (bank["Ledger"].str.upper() != "UNMAPPED")][["Ledger", amt_col]].copy()
        dr_list = dr_list.rename(columns={amt_col: "Amount"})

        mapped_cr_total = cr_list["Amount"].sum() if len(cr_list) else 0.0
        mapped_dr_total = dr_list["Amount"].sum() if len(dr_list) else 0.0
        receipt_diff = round(total_cr - mapped_cr_total, 2)
        payment_diff = round(total_dr - mapped_dr_total, 2)

        if abs(receipt_diff) > 0.01:
            cr_list = pd.concat([cr_list, pd.DataFrame([{"Ledger": "Others", "Amount": receipt_diff}])], ignore_index=True)
        if abs(payment_diff) > 0.01:
            dr_list = pd.concat([dr_list, pd.DataFrame([{"Ledger": "Others", "Amount": payment_diff}])], ignore_index=True)

        cr_list = move_others_last_df(cr_list)
        dr_list = move_others_last_df(dr_list)
        daybook_df = make_daybook_layout_from_lists(cr_list, dr_list)

    # Mapping DF
    stmt_cols = []
    if "Tran Date" in bank.columns:
        stmt_cols.append("Tran Date")
    if "Value Date" in bank.columns:
        stmt_cols.append("Value Date")
    if "Txn Posted Date" in bank.columns:
        stmt_cols.append("Txn Posted Date")
    stmt_cols += [desc_col, crdr_col]

    mapping_df = bank[stmt_cols + [amt_col, "Ledger"]].copy()
    mapping_df = mapping_df.rename(columns={amt_col: "Amount"})

    # Entry DF
    doc_date = pick_document_date(bank)
    entry_df = create_entry_df(daybook_df, display_name, coa_lookup_by_name, coa_name_by_pair, doc_date)

    # Styles
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="BFBFBF")
    amt_fill = PatternFill("solid", fgColor="FFF200")
    others_fill = PatternFill("solid", fgColor="F8CBAD")
    unmapped_fill = PatternFill("solid", fgColor="FFC7CE")
    bold = Font(bold=True)

    db_name = safe_sheet_name(f"DayBook_{display_name}")
    mp_name = safe_sheet_name(f"Mapping_{display_name}")
    en_name = safe_sheet_name(f"Entry_{display_name}")

    def unique_name(nm):
        if nm not in wb.sheetnames:
            return nm
        i = 1
        while True:
            nn = safe_sheet_name(f"{nm}_{i}")
            if nn not in wb.sheetnames:
                return nn
            i += 1

    db_name = unique_name(db_name)
    mp_name = unique_name(mp_name)
    en_name = unique_name(en_name)

    # DAYBOOK
    ws = wb.create_sheet(db_name)
    ws["A1"] = title_text
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = Alignment(horizontal="center")

    start_row = 3
    headers = ["RECEIPT", "AMOUNT", "PAYMENTS", "AMOUNT"]
    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    write_df = daybook_df.copy()
    if write_df.columns.tolist() == ["RECEIPT", "AMOUNT", "PAYMENTS", "AMOUNT"]:
        write_df.columns = ["RECEIPT", "AMOUNT", "PAYMENTS", "AMOUNT.1"]

    mp_link_sheet = excel_hyperlink_sheet(mp_name)

    for r in range(len(write_df)):
        rec_led = clean_text(write_df.iat[r, 0]).upper()
        pay_led = clean_text(write_df.iat[r, 2]).upper()
        row_vals = [write_df.iat[r, 0], write_df.iat[r, 1], write_df.iat[r, 2], write_df.iat[r, 3]]
        for c, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=start_row + 1 + r, column=c, value=val)
            if c in [2, 4]:
                cell.fill = amt_fill
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            if c in [1, 2] and rec_led == "OTHERS":
                cell.fill = others_fill
                if c == 1:
                    cell.hyperlink = f"#'{mp_link_sheet}'!A1"
                    cell.font = Font(color="0563C1", underline="single")

            if c in [3, 4] and pay_led == "OTHERS":
                cell.fill = others_fill
                if c == 3:
                    cell.hyperlink = f"#'{mp_link_sheet}'!A1"
                    cell.font = Font(color="0563C1", underline="single")

            cell.border = border

    last_data_row = start_row + len(write_df)
    total_row = last_data_row + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = bold
    ws.cell(row=total_row, column=1).fill = header_fill
    ws.cell(row=total_row, column=1).border = border
    ws.cell(row=total_row, column=2, value=f"=SUM(B{start_row+1}:B{last_data_row})").font = bold
    ws.cell(row=total_row, column=2).fill = amt_fill
    ws.cell(row=total_row, column=2).number_format = "#,##0.00"
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=2).border = border
    ws.cell(row=total_row, column=3, value="TOTAL").font = bold
    ws.cell(row=total_row, column=3).fill = header_fill
    ws.cell(row=total_row, column=3).border = border
    ws.cell(row=total_row, column=4, value=f"=SUM(D{start_row+1}:D{last_data_row})").font = bold
    ws.cell(row=total_row, column=4).fill = amt_fill
    ws.cell(row=total_row, column=4).number_format = "#,##0.00"
    ws.cell(row=total_row, column=4).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=4).border = border

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 18

    # MAPPING (hide mapped rows)
    ws2 = wb.create_sheet(mp_name)
    for col_idx, col_name in enumerate(mapping_df.columns, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=col_name)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for r in range(len(mapping_df)):
        ledger_val = str(mapping_df.loc[r, "Ledger"]).strip().upper()
        is_unmapped = ledger_val == "UNMAPPED"
        excel_row = r + 2
        if not is_unmapped:
            ws2.row_dimensions[excel_row].hidden = True
        for c, col_name in enumerate(mapping_df.columns, start=1):
            val = mapping_df.loc[r, col_name]
            cell = ws2.cell(row=excel_row, column=c, value=val)
            if col_name.upper() == "AMOUNT":
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            if is_unmapped:
                cell.fill = unmapped_fill
            cell.border = border

    last_row = len(mapping_df) + 1
    last_col = len(mapping_df.columns)
    last_col_letter = chr(64 + last_col)
    table_ref = f"A1:{last_col_letter}{last_row}"
    tbl_name = f"Tbl_{abs(hash(mp_name)) % 10_000_000}"
    table = Table(displayName=tbl_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False
    )
    ws2.add_table(table)
    ws2.auto_filter.ref = table_ref
    ws2.freeze_panes = "A2"

    # ENTRY
    ws3 = wb.create_sheet(en_name)
    for col_idx, col_name in enumerate(entry_df.columns, start=1):
        cell = ws3.cell(row=1, column=col_idx, value=col_name)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for r in range(len(entry_df)):
        excel_row = r + 2
        for c, col_name in enumerate(entry_df.columns, start=1):
            val = entry_df.iloc[r][col_name]
            cell = ws3.cell(row=excel_row, column=c, value=val)
            if col_name in ["Debit", "Credit"] and val != "" and val is not None:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border

    ws3.freeze_panes = "A2"

# =========================
# ✅ STREAMLIT ENTRYPOINT
# =========================
def run_daybook_from_uploaded_files(coa_bytes: bytes, zip_bytes: bytes) -> bytes:
    """
    Streamlit-friendly runner:
    - receives COA.xlsx bytes and Statement.zip bytes
    - returns HO_DayBook_AllStatements.xlsx bytes
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        coa_path = os.path.join(tmpdir, "COA.xlsx")
        zip_path = os.path.join(tmpdir, "Statement.zip")
        out_path = os.path.join(tmpdir, "HO_DayBook_AllStatements.xlsx")

        with open(coa_path, "wb") as f:
            f.write(coa_bytes)
        with open(zip_path, "wb") as f:
            f.write(zip_bytes)

        coa = pd.read_excel(coa_path, sheet_name=0)
        kw_df = build_keyword_table(coa)
        coa_lookup_by_name, coa_name_by_pair = build_coa_lookup(coa)

        wb = Workbook()
        wb.remove(wb.active)

        extract_dir = os.path.join(tmpdir, "unzipped")
        os.makedirs(extract_dir, exist_ok=True)
        with zipfile.ZipFile(zip_path, "r") as z:
            z.extractall(extract_dir)

        statement_files = []
        for root, _, files in os.walk(extract_dir):
            for f in files:
                if f.lower().endswith(".xlsx") and not f.startswith("~$"):
                    statement_files.append(os.path.join(root, f))

        if not statement_files:
            raise ValueError("No .xlsx bank statements found inside the ZIP.")

        errors = []
        for fp in sorted(statement_files):
            try:
                add_all_sheets_for_statement(wb, fp, kw_df, coa_lookup_by_name, coa_name_by_pair)
            except Exception as e:
                errors.append(f"{os.path.basename(fp)} -> {e}")

        wb.save(out_path)

        # If you want to show errors in UI, we can return them separately later.
        with open(out_path, "rb") as f:
            output_bytes = f.read()

        return output_bytes

# app.py
# Streamlit app: Collection Efficiency Automation (DVC consolidate + Hub-wise split)
# + "Arrear_Advance" pivot sheet (cust_id wise) in the DVC output.
# + NEW MODULE: "LPC Report" (filtered detail + formatted Summary)
# IMPORTANT:
# - Arrear hub-wise saved as CSV (to avoid MemoryError).
# - LPC detail sheet is written by pandas (fast, low memory) + Summary formatted by openpyxl.

import os
import re
import zipfile
import tempfile
import time
import pandas as pd
import numpy as np
import streamlit as st

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import re
from collections import OrderedDict, defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------
# Vault Cash Logic (Single-file version)
# ----------------------------
DEST_SHEET = "Filtered Sheet"
ESC_SHEET = "Escalation"
SUMMARY_SHEET = "Summary"

NEEDED_VAULT = ["Branch", "Cash Date", "Vault Balance", "System Balance", "Cash Difference", "Modified By"]


def find_header_row(ws, must_have=("Branch",), scan_rows=20):
    for r in range(1, scan_rows + 1):
        row_vals = [str(c.value).strip() if c.value is not None else "" for c in ws[r]]
        if all(any(v.lower() == h.lower() for v in row_vals) for h in must_have):
            return r
    raise ValueError(f"Header row not found in sheet '{ws.title}' (scanned first {scan_rows} rows).")


def build_col_map(ws, header_row):
    col_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        key = str(v).strip()
        if key:
            col_map[key.lower()] = c
    return col_map


def split_branch(branch_text):
    if branch_text is None:
        return "", ""
    s = str(branch_text).strip()
    s = re.sub(r"\s*-\s*", "-", s)
    if "-" in s:
        code, name = s.split("-", 1)
        return code.strip(), name.strip()
    return s, ""


def norm_code(x):
    if x is None:
        return ""
    return str(x).strip().upper()


def is_blank(v):
    return v is None or str(v).strip() == ""


def clear_below(ws, start_row, start_col=1, end_col=None):
    if end_col is None:
        end_col = ws.max_column
    for r in range(start_row, ws.max_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).value = None


def apply_table_format(ws, min_row, max_row, min_col, max_col, header_row=None):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(r, c)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    if header_row:
        header_fill = PatternFill("solid", fgColor="BDD7EE")
        header_font = Font(bold=True)
        for c in range(min_col, max_col + 1):
            cell = ws.cell(header_row, c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")


def autofit_columns(ws, from_col, to_col, min_width=10, max_width=45):
    for c in range(from_col, to_col + 1):
        max_len = 0
        col_letter = get_column_letter(c)
        for cell in ws[col_letter]:
            if cell.value is None:
                continue
            v = str(cell.value)
            max_len = max(max_len, len(v))
        width = max(min_width, min(max_width, max_len + 2))
        ws.column_dimensions[col_letter].width = width


def build_summary_from_filtered(wb, ws_filtered, dst_cols, data_start_row, data_end_row):
    if SUMMARY_SHEET in wb.sheetnames:
        ws_sum = wb[SUMMARY_SHEET]
        wb.remove(ws_sum)
    ws_sum = wb.create_sheet(SUMMARY_SHEET)

    zone_order = OrderedDict()
    counts = defaultdict(int)

    for r in range(data_start_row, data_end_row + 1):
        zone = ws_filtered.cell(r, dst_cols["zone"]).value
        region = ws_filtered.cell(r, dst_cols["region"]).value
        remarks = ws_filtered.cell(r, dst_cols["remarks"]).value

        if is_blank(zone) or is_blank(region) or is_blank(remarks):
            continue

        z = str(zone).strip()
        reg = str(region).strip()
        rem = str(remarks).strip()

        if z not in zone_order:
            zone_order[z] = True

        counts[(z, reg, rem)] += 1

    headers = ["Zone", "Region", "Filled by branch", "Not Filled by Branch", "Count of Branches"]
    ws_sum.append(headers)

    block_fills = [
        PatternFill("solid", fgColor="F8CBAD"),
        PatternFill("solid", fgColor="D9D9D9"),
        PatternFill("solid", fgColor="FFE699"),
        PatternFill("solid", fgColor="C6E0B4"),
    ]
    total_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    current_row = 2
    grand_filled = 0
    grand_not = 0
    grand_total = 0

    def regions_for_zone(z):
        regs = set()
        for (zz, rr, _rem) in counts.keys():
            if zz == z:
                regs.add(rr)
        return sorted(regs)

    zone_index = 0
    for z in zone_order.keys():
        fill = block_fills[zone_index % len(block_fills)]
        zone_index += 1

        z_filled_sum = 0
        z_not_sum = 0
        z_total_sum = 0

        for reg in regions_for_zone(z):
            filled_cnt = counts.get((z, reg, "Filled by Branch"), 0)
            not_cnt = counts.get((z, reg, "Not Filled by Branch"), 0)
            total_cnt = filled_cnt + not_cnt

            ws_sum.cell(current_row, 1).value = z
            ws_sum.cell(current_row, 2).value = reg
            ws_sum.cell(current_row, 3).value = filled_cnt
            ws_sum.cell(current_row, 4).value = not_cnt if not_cnt != 0 else "-"
            ws_sum.cell(current_row, 5).value = total_cnt

            for c in range(1, 6):
                cell = ws_sum.cell(current_row, c)
                cell.fill = fill
                cell.alignment = center if c >= 3 else left

            z_filled_sum += filled_cnt
            z_not_sum += not_cnt
            z_total_sum += total_cnt
            current_row += 1

        ws_sum.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        ws_sum.cell(current_row, 1).value = f"{z} Total"
        ws_sum.cell(current_row, 3).value = z_filled_sum
        ws_sum.cell(current_row, 4).value = z_not_sum if z_not_sum != 0 else "-"
        ws_sum.cell(current_row, 5).value = z_total_sum

        for c in range(1, 6):
            cell = ws_sum.cell(current_row, c)
            cell.fill = fill
            cell.font = total_font
            cell.alignment = center if c >= 3 else left

        current_row += 1
        grand_filled += z_filled_sum
        grand_not += z_not_sum
        grand_total += z_total_sum

    grand_fill = PatternFill("solid", fgColor="9BC2E6")
    ws_sum.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    ws_sum.cell(current_row, 1).value = "Grand Total"
    ws_sum.cell(current_row, 3).value = grand_filled
    ws_sum.cell(current_row, 4).value = grand_not if grand_not != 0 else "-"
    ws_sum.cell(current_row, 5).value = grand_total

    for c in range(1, 6):
        cell = ws_sum.cell(current_row, c)
        cell.fill = grand_fill
        cell.font = Font(bold=True, color="000000")
        cell.alignment = center if c >= 3 else left

    apply_table_format(ws_sum, min_row=1, max_row=current_row, min_col=1, max_col=5, header_row=1)

    for r in range(2, current_row + 1):
        ws_sum.cell(r, 1).alignment = left
        ws_sum.cell(r, 2).alignment = left
        for c in (3, 4, 5):
            ws_sum.cell(r, c).alignment = center

    ws_sum.freeze_panes = "A2"
    autofit_columns(ws_sum, 1, 5)
    return ws_sum


def run_vault_cash_report(file_path: str):
    wb = load_workbook(file_path)

    # ✅ AUTO FIX: choose "Vault" if "Valut" not present
    possible_src = ["Valut", "Vault"]
    SRC_SHEET = next((s for s in possible_src if s in wb.sheetnames), None)
    if SRC_SHEET is None:
        raise ValueError(f"Vault sheet not found. Expected one of {possible_src}. Found: {wb.sheetnames}")

    for sh in (SRC_SHEET, DEST_SHEET, ESC_SHEET):
        if sh not in wb.sheetnames:
            raise ValueError(f"Sheet '{sh}' not found. Available: {wb.sheetnames}")

    ws_src = wb[SRC_SHEET]
    ws_dst = wb[DEST_SHEET]
    ws_esc = wb[ESC_SHEET]

    esc_header_row = find_header_row(ws_esc, must_have=("Branch Code",))
    esc_cols = build_col_map(ws_esc, esc_header_row)

    required_esc = ["branch code", "branch name", "zone", "region"]
    miss_esc = [x for x in required_esc if x not in esc_cols]
    if miss_esc:
        raise ValueError(f"Missing these headers in '{ESC_SHEET}' sheet: {miss_esc}")

    esc_map = {}
    for r in range(esc_header_row + 1, ws_esc.max_row + 1):
        code = norm_code(ws_esc.cell(r, esc_cols["branch code"]).value)
        if not code:
            continue
        if code not in esc_map:
            esc_map[code] = {
                "zone": ws_esc.cell(r, esc_cols["zone"]).value,
                "region": ws_esc.cell(r, esc_cols["region"]).value,
                "branch_name": ws_esc.cell(r, esc_cols["branch name"]).value,
            }

    dst_header_row = find_header_row(ws_dst, must_have=("Code", "Branch"))
    dst_cols = build_col_map(ws_dst, dst_header_row)

    dst_needed = ["Zone", "Region", "Code", "Branch", "Cash Date", "Vault Balance",
                  "System Balance", "Cash Difference", "Modified By", "Remarks"]
    dst_missing = [h for h in dst_needed if h.lower() not in dst_cols]
    if dst_missing:
        raise ValueError(f"Missing these headers in '{DEST_SHEET}': {dst_missing}")

    src_header_row = find_header_row(ws_src, must_have=("Branch",))
    src_cols = build_col_map(ws_src, src_header_row)

    miss_vault = [h for h in NEEDED_VAULT if h.lower() not in src_cols]
    if miss_vault:
        raise ValueError(f"Missing these headers in '{SRC_SHEET}': {miss_vault}")
    if "created by" not in src_cols:
        raise ValueError(f"Missing 'Created By' header in '{SRC_SHEET}' sheet (needed to fill Modified By).")

    dst_write_start = 2
    clear_below(ws_dst, start_row=dst_write_start)

    out_r = dst_write_start
    seen_codes_vault = set()

    for r in range(src_header_row + 1, ws_src.max_row + 1):
        row_is_blank = True
        for h in NEEDED_VAULT:
            if ws_src.cell(r, src_cols[h.lower()]).value not in (None, ""):
                row_is_blank = False
                break
        if row_is_blank:
            continue

        branch_val = ws_src.cell(r, src_cols["branch"]).value
        code, branch_name_from_vault = split_branch(branch_val)
        code_n = norm_code(code)
        if not code_n:
            continue

        if code_n in seen_codes_vault:
            continue
        seen_codes_vault.add(code_n)

        esc = esc_map.get(code_n, {})
        zone_val = esc.get("zone")
        region_val = esc.get("region")

        branch_name = esc.get("branch_name")
        if is_blank(branch_name):
            branch_name = branch_name_from_vault

        ws_dst.cell(out_r, dst_cols["zone"]).value = zone_val
        ws_dst.cell(out_r, dst_cols["region"]).value = region_val
        ws_dst.cell(out_r, dst_cols["code"]).value = code
        ws_dst.cell(out_r, dst_cols["branch"]).value = branch_name

        for col_name in ["Cash Date", "Vault Balance", "System Balance", "Cash Difference"]:
            ws_dst.cell(out_r, dst_cols[col_name.lower()]).value = ws_src.cell(r, src_cols[col_name.lower()]).value

        mod_by = ws_src.cell(r, src_cols["modified by"]).value
        if is_blank(mod_by):
            mod_by = ws_src.cell(r, src_cols["created by"]).value
        ws_dst.cell(out_r, dst_cols["modified by"]).value = mod_by

        ws_dst.cell(out_r, dst_cols["remarks"]).value = "Filled by Branch"
        out_r += 1

    present_codes = set(seen_codes_vault)
    for code_n, info in esc_map.items():
        if code_n in present_codes:
            continue
        ws_dst.cell(out_r, dst_cols["zone"]).value = info.get("zone")
        ws_dst.cell(out_r, dst_cols["region"]).value = info.get("region")
        ws_dst.cell(out_r, dst_cols["code"]).value = code_n
        ws_dst.cell(out_r, dst_cols["branch"]).value = info.get("branch_name")
        ws_dst.cell(out_r, dst_cols["remarks"]).value = "Not Filled by Branch"
        out_r += 1

    last_data_row = out_r - 1

    build_summary_from_filtered(
        wb=wb,
        ws_filtered=ws_dst,
        dst_cols=dst_cols,
        data_start_row=dst_write_start,
        data_end_row=last_data_row
    )

    wb.save(file_path)
NotImplementedError("Vault Cash logic not pasted yet")

def run_pending_collection_entry(excel_path):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # -----------------------------
    # 1) Read Details + Apply Filters
    # -----------------------------
    df = pd.read_excel(excel_path, sheet_name="Details")
    df.columns = [str(c).strip() for c in df.columns]

    required_cols = ["Sub Account", "Opening Credit", "Closing Credit", "Moving Debit", "Moving Credit"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Missing column(s) in Details sheet: {missing}")

    df_f = df.copy()
    df_f = df_f[df_f["Closing Credit"].fillna(0) != 0]   # Closing Credit ≠ 0
    df_f = df_f[df_f["Opening Credit"].fillna(0) != 0]   # Opening Credit ≠ 0
    df_f = df_f[df_f["Moving Debit"].fillna(0) == 0]     # Moving Debit = 0
    df_f = df_f[df_f["Moving Credit"].fillna(0) == 0]    # Moving Credit = 0

    out_df = df_f[["Sub Account", "Opening Credit", "Closing Credit"]].copy()

    # -----------------------------
    # 2) Read Escalation + Lookup Maker Name
    # -----------------------------
    esc = pd.read_excel(excel_path, sheet_name="Escalation")
    esc.columns = [str(c).strip() for c in esc.columns]

    def norm(s):
        return "".join(str(s).strip().lower().split())

    norm_cols = {norm(c): c for c in esc.columns}

    branch_candidates = ["branchcode", "branch_code", "branch", "subaccount", "sub_account"]
    branch_col = None
    for key in branch_candidates:
        if key in norm_cols:
            branch_col = norm_cols[key]
            break
    if branch_col is None:
        raise ValueError(f"Could not find Branch Code column in Escalation. Columns: {list(esc.columns)}")

    maker_candidates = ["makername", "maker_name", "maker", "makerusername", "makerid", "maker_id"]
    maker_col = None
    for key in [norm(x) for x in maker_candidates]:
        if key in norm_cols:
            maker_col = norm_cols[key]
            break
    if maker_col is None:
        raise ValueError(f"Could not find Maker Name column in Escalation. Columns: {list(esc.columns)}")

    esc_map = esc[[branch_col, maker_col]].dropna(subset=[branch_col]).copy()
    esc_map[branch_col] = (
        esc_map[branch_col]
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
    )
    esc_map[maker_col] = esc_map[maker_col].astype(str).str.strip()

    maker_lookup = dict(zip(esc_map[branch_col], esc_map[maker_col]))

    out_df["Maker name"] = (
        out_df["Sub Account"]
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
        .map(maker_lookup)
        .fillna("")
    )

    out_df = out_df.rename(columns={"Sub Account": "Branch Code"})
    out_df = out_df[["Maker name", "Branch Code", "Opening Credit", "Closing Credit"]]

    # -----------------------------
    # 3) Write to "summary" sheet
    # -----------------------------
    wb = load_workbook(excel_path)
    if "summary" in wb.sheetnames:
        wb.remove(wb["summary"])
    ws = wb.create_sheet("summary")

    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    for c, name in enumerate(out_df.columns, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    # Data
    for r, row in enumerate(out_df.itertuples(index=False), start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            if c in (3, 4):
                cell.number_format = "#,##0.00"

    ws.freeze_panes = "A2"

    # Auto width
    for c in range(1, len(out_df.columns) + 1):
        col = get_column_letter(c)
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in ws[col])
        ws.column_dimensions[col].width = min(max_len + 3, 40)

    wb.save(excel_path)


import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SHEET_DVC_UPDATED = "DVC_Updated"
SHEET_DVC_PIVOT = "DVC_Pivot"
SHEET_DEMCOL_PIVOT = "DEMCOL_Pivot"


def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(c).strip().upper() for c in df.columns]
    return df


def norm_loan_id(x):
    if pd.isna(x):
        return ""
    s = str(x).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def insert_due_demand(dvc: pd.DataFrame) -> pd.DataFrame:
    dvc = clean_columns(dvc)

    required = {"CURRENT_PRINCIPAL_DEMAND", "CURRENT_INTEREST_DEMAND"}
    missing = required - set(dvc.columns)
    if missing:
        raise KeyError(f"DVC missing columns after cleaning: {missing}\nColumns found: {list(dvc.columns)}")

    dvc["CURRENT_PRINCIPAL_DEMAND"] = pd.to_numeric(dvc["CURRENT_PRINCIPAL_DEMAND"], errors="coerce").fillna(0)
    dvc["CURRENT_INTEREST_DEMAND"] = pd.to_numeric(dvc["CURRENT_INTEREST_DEMAND"], errors="coerce").fillna(0)

    due = dvc["CURRENT_PRINCIPAL_DEMAND"] + dvc["CURRENT_INTEREST_DEMAND"]

    cols = list(dvc.columns)
    if "DUE DEMAND" in cols:
        dvc["DUE DEMAND"] = due
        return dvc

    idx = cols.index("CURRENT_INTEREST_DEMAND") + 1
    cols = cols[:idx] + ["DUE DEMAND"] + cols[idx:]
    dvc["DUE DEMAND"] = due
    return dvc[cols]


def pivot_dvc(dvc: pd.DataFrame) -> pd.DataFrame:
    dvc = clean_columns(dvc)

    if "LOAN_ID" not in dvc.columns:
        raise KeyError(f"DVC missing column LOAN_ID. Found: {list(dvc.columns)}")

    p = (dvc.pivot_table(index="LOAN_ID", values="DUE DEMAND", aggfunc="sum", fill_value=0)
         .reset_index()
         .rename(columns={"DUE DEMAND": "DVC_DUE_DEMAND"}))

    p["LOAN_ID"] = p["LOAN_ID"].apply(norm_loan_id)
    return p.sort_values("LOAN_ID")


def pivot_demcol(dem: pd.DataFrame) -> pd.DataFrame:
    dem = clean_columns(dem)

    required = {"LOAN_ID", "TOTAL_AMT_DUE"}
    missing = required - set(dem.columns)
    if missing:
        raise KeyError(f"DEMCOL missing columns after cleaning: {missing}\nColumns found: {list(dem.columns)}")

    dem["TOTAL_AMT_DUE"] = pd.to_numeric(dem["TOTAL_AMT_DUE"], errors="coerce").fillna(0)

    p = (dem.pivot_table(index="LOAN_ID", values="TOTAL_AMT_DUE", aggfunc="sum", fill_value=0)
         .reset_index()
         .rename(columns={"TOTAL_AMT_DUE": "DEMCOL_TOTAL_AMT_DUE"}))

    p["LOAN_ID"] = p["LOAN_ID"].apply(norm_loan_id)
    return p.sort_values("LOAN_ID")


def write_df(wb, sheet_name: str, df: pd.DataFrame):
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)

    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))

    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = len(str(col_name))
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    ws.freeze_panes = "A2"


def build_lookup(df: pd.DataFrame, key_col: str, pick_cols: list) -> pd.DataFrame:
    df = clean_columns(df)

    need = set([key_col] + pick_cols)
    missing = need - set(df.columns)
    if missing:
        raise KeyError(f"Missing in source sheet for lookup: {missing}\nFound: {list(df.columns)}")

    tmp = df[[key_col] + pick_cols].copy()
    tmp[key_col] = tmp[key_col].apply(norm_loan_id)

    tmp = tmp.drop_duplicates(subset=[key_col], keep="first")
    return tmp.set_index(key_col)


def add_status_columns_for_diff(dvc_compare: pd.DataFrame,
                                dem_compare: pd.DataFrame,
                                dvc_raw: pd.DataFrame,
                                dem_raw: pd.DataFrame):

    dvc_lu = build_lookup(
        dvc_raw,
        key_col="LOAN_ID",
        pick_cols=["STATUS", "PROD_CATEGORY_ID", "LOAN_MATURITY_DATE"]
    )

    dem_lu = build_lookup(
        dem_raw,
        key_col="LOAN_ID",
        pick_cols=["STATUS", "PRODUCT_ID", "CLOSURE_DATE"]
    )

    dvc_out = dvc_compare.copy()
    dvc_out["LOAN_ID"] = dvc_out["LOAN_ID"].apply(norm_loan_id)
    is_diff = dvc_out["DIFFERENCE (DVC - DEMCOL)"].fillna(0) != 0

    dvc_out["WRITE OFF"] = ""
    dvc_out["PRODUCT"] = ""
    dvc_out["MATURED"] = ""

    idxs = dvc_out.loc[is_diff, "LOAN_ID"]
    dvc_out.loc[is_diff, "WRITE OFF"] = idxs.map(dvc_lu["STATUS"]).fillna("")
    dvc_out.loc[is_diff, "PRODUCT"] = idxs.map(dvc_lu["PROD_CATEGORY_ID"]).fillna("")
    dvc_out.loc[is_diff, "MATURED"] = idxs.map(dvc_lu["LOAN_MATURITY_DATE"]).fillna("")

    dem_out = dem_compare.copy()
    dem_out["LOAN_ID"] = dem_out["LOAN_ID"].apply(norm_loan_id)
    is_diff2 = dem_out["DIFFERENCE (DEMCOL - DVC)"].fillna(0) != 0

    dem_out["WRITE OFF"] = ""
    dem_out["PRODUCT"] = ""
    dem_out["MATURED"] = ""

    idxs2 = dem_out.loc[is_diff2, "LOAN_ID"]
    dem_out.loc[is_diff2, "WRITE OFF"] = idxs2.map(dem_lu["STATUS"]).fillna("")
    dem_out.loc[is_diff2, "PRODUCT"] = idxs2.map(dem_lu["PRODUCT_ID"]).fillna("")
    dem_out.loc[is_diff2, "MATURED"] = idxs2.map(dem_lu["CLOSURE_DATE"]).fillna("")

    return dvc_out, dem_out


def run_demand_verification(file_path: str):
    if not os.path.exists(file_path):
        raise FileNotFoundError(file_path)

    dvc = pd.read_excel(file_path, sheet_name="DVC")
    demcol = pd.read_excel(file_path, sheet_name="DEMCOL")

    dvc_updated = insert_due_demand(dvc)

    dvc_p = pivot_dvc(dvc_updated)
    dem_p = pivot_demcol(demcol)

    dvc_compare = dvc_p.merge(dem_p, on="LOAN_ID", how="left")
    dvc_compare["DEMCOL_TOTAL_AMT_DUE"] = dvc_compare["DEMCOL_TOTAL_AMT_DUE"].fillna(0)
    dvc_compare["DIFFERENCE (DVC - DEMCOL)"] = dvc_compare["DVC_DUE_DEMAND"] - dvc_compare["DEMCOL_TOTAL_AMT_DUE"]

    dem_compare = dem_p.merge(dvc_p, on="LOAN_ID", how="left")
    dem_compare["DVC_DUE_DEMAND"] = dem_compare["DVC_DUE_DEMAND"].fillna(0)
    dem_compare["DIFFERENCE (DEMCOL - DVC)"] = dem_compare["DEMCOL_TOTAL_AMT_DUE"] - dem_compare["DVC_DUE_DEMAND"]

    # NEW status columns
    dvc_compare, dem_compare = add_status_columns_for_diff(dvc_compare, dem_compare, dvc, demcol)

    wb = load_workbook(file_path)
    write_df(wb, SHEET_DVC_UPDATED, dvc_updated)
    write_df(wb, SHEET_DVC_PIVOT, dvc_compare)
    write_df(wb, SHEET_DEMCOL_PIVOT, dem_compare)
    wb.save(file_path)

import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

TOL = 0.000001


def to_num(x):
    try:
        if x is None or x == "":
            return 0.0
        return float(x)
    except:
        return 0.0


def norm(s):
    return str(s).strip().lower() if s is not None else ""


def is_branch_id(val):
    s = str(val).strip() if val is not None else ""
    return bool(re.match(r"^B\d+", s, re.IGNORECASE))


# ✅ STREAMLIT CLOUD SAFE: Excel recalculation not available
def force_excel_recalc_and_save(_path: str):
    # openpyxl cannot calculate formulas and Streamlit Cloud has no Excel engine.
    # So we skip recalculation here.
    return


def create_consolidated_sheet(map_file_path):
    """
    Creates/updates 'Consolidated' sheet in MAP.xlsx by taking only rows where:
      - Column A = 'Closing Balance'
      - Column B starts with 'B' (Branch ID)
    Copies columns A to I (A..I, where I is Balance Amount)
    """
    CONSOL_SHEET = "Consolidated"

    # ✅ skip excel recalculation on cloud
    force_excel_recalc_and_save(map_file_path)

    COL_REGION = "A"
    COL_BRANCH = "B"
    COPY_COLS = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]

    # Read cached/calculated values if present
    wb_val = load_workbook(map_file_path, data_only=True)
    wb = load_workbook(map_file_path)

    if CONSOL_SHEET in wb.sheetnames:
        del wb[CONSOL_SHEET]

    ws_out = wb.create_sheet(CONSOL_SHEET, 0)
    ws_out.append(["Source Sheet"] + COPY_COLS)

    for sh in wb.sheetnames:
        if sh == CONSOL_SHEET:
            continue

        ws = wb[sh]
        ws_v = wb_val[sh]

        for r in range(1, ws.max_row + 1):
            region_val = ws[f"{COL_REGION}{r}"].value
            branch_val = ws[f"{COL_BRANCH}{r}"].value

            if norm(region_val) == "closing balance" and is_branch_id(branch_val):
                row_data = [sh]
                for col in COPY_COLS:
                    row_data.append(ws_v[f"{col}{r}"].value)
                ws_out.append(row_data)

    wb.save(map_file_path)


def format_difference_sheet(ws):
    header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    bold_font = Font(bold=True)

    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = header_fill

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    max_col = ws.max_column

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).border = border

    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        max_len = 0
        for r in range(1, max_row + 1):
            v = ws.cell(r, c).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = max_len + 3

    for r in range(1, max_row + 1):
        ws.row_dimensions[r].height = 20

    ws.freeze_panes = "A2"


def run_map_ledger_difference(map_file_path, ledger_file_path):
    CONSOL_SHEET = "Consolidated"
    LEDGER_SHEET = "Ledger"
    ESC_SHEET = "Escalation"
    DIFF_SHEET = "Difference"

    # Step 0: Create/Update Consolidated in MAP.xlsx (NO EXCEL RECALC ON CLOUD)
    create_consolidated_sheet(map_file_path)

    # 1) Read MAP Consolidated
    df_map = pd.read_excel(map_file_path, sheet_name=CONSOL_SHEET, engine="openpyxl")
    branch_col = "B"
    balance_col = "I"

    df_map[branch_col] = df_map[branch_col].astype(str).str.strip()
    df_map[balance_col] = pd.to_numeric(df_map[balance_col], errors="coerce").fillna(0)
    map_balance_by_branch = df_map.set_index(branch_col)[balance_col].to_dict()

    # 2) Open MAP Ledger workbook
    wb = load_workbook(ledger_file_path)
    ws_ledger = wb[LEDGER_SHEET]
    ws_esc = wb[ESC_SHEET]

    # 3) Escalation mapping Branch Code -> Maker (Executive Name)
    esc_headers = {}
    for c in range(1, ws_esc.max_column + 1):
        h = ws_esc.cell(1, c).value
        if h:
            esc_headers[norm(h)] = c

    col_branch_code = esc_headers.get("branch code")
    col_maker = esc_headers.get("maker")
    if not col_branch_code or not col_maker:
        raise Exception("Escalation sheet must have headers: 'Branch Code' and 'Maker' in row 1")

    exec_by_branch = {}
    for r in range(2, ws_esc.max_row + 1):
        bc = ws_esc.cell(r, col_branch_code).value
        mk = ws_esc.cell(r, col_maker).value
        if bc is None:
            continue
        exec_by_branch[str(bc).strip()] = "" if mk is None else str(mk).strip()

    # 4) Ledger columns
    ledger_headers = {}
    for c in range(1, ws_ledger.max_column + 1):
        h = ws_ledger.cell(1, c).value
        if h:
            ledger_headers[norm(h)] = c

    sub_acc_col = ledger_headers.get("sub account")
    debit_col = ledger_headers.get("closing debit")
    credit_col = ledger_headers.get("closing credit")
    if not (sub_acc_col and debit_col and credit_col):
        raise Exception("Ledger sheet must have headers: 'Sub Account', 'Closing Debit', 'Closing Credit' in row 1")

    # 5) Create/Replace Difference sheet
    if DIFF_SHEET in wb.sheetnames:
        del wb[DIFF_SHEET]
    ws_diff = wb.create_sheet(DIFF_SHEET)

    ws_diff.append([
        "Executive Name",
        "Branch Code",
        "Ledger Balance (Cl.Credit-Cl.Debit)",
        "MAP Sheet Balance",
        "Difference"
    ])

    # 6) Only difference rows
    for r in range(2, ws_ledger.max_row + 1):
        sub_acc = ws_ledger.cell(r, sub_acc_col).value
        if sub_acc is None or str(sub_acc).strip() == "":
            continue

        key = str(sub_acc).strip()

        closing_credit = to_num(ws_ledger.cell(r, credit_col).value)
        closing_debit = to_num(ws_ledger.cell(r, debit_col).value)
        final_balance = closing_credit - closing_debit

        map_balance = to_num(map_balance_by_branch.get(key, 0))
        diff = final_balance - map_balance

        if abs(diff) > TOL:
            exec_name = exec_by_branch.get(key, "")
            ws_diff.append([exec_name, key, final_balance, map_balance, diff])

    format_difference_sheet(ws_diff)
    wb.save(ledger_file_path)

import os
import pandas as pd
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_SHEET_NAME = "Excess_Amount_Received"
CONSOLIDATED_SHEET = "Consolidated"
ESCALATION_SHEET = "Sheet1"


def norm_col(s: str) -> str:
    return (
        str(s).strip().upper()
        .replace(" ", "")
        .replace("-", "")
        .replace(".", "")
        .replace("\n", "")
        .replace("\t", "")
    )


def find_col(df, required_name):
    req = norm_col(required_name)
    col_map = {norm_col(c): c for c in df.columns}
    return col_map.get(req)


def to_num(series):
    return pd.to_numeric(series, errors="coerce").fillna(0)


def read_sheet_with_status(file_path: str):
    xl = pd.ExcelFile(file_path, engine="openpyxl")
    for sh in xl.sheet_names:
        df = xl.parse(sh)
        if find_col(df, "Status") is not None:
            return df, sh
    df0 = xl.parse(xl.sheet_names[0])
    return df0, xl.sheet_names[0]


def load_escalation_map(escalation_path: str):
    esc = pd.read_excel(escalation_path, sheet_name=ESCALATION_SHEET, engine="openpyxl")

    col_bc = find_col(esc, "Branch Code")
    col_maker = find_col(esc, "Maker")

    if col_bc is None or col_maker is None:
        raise ValueError("Escalation.xlsx must contain 'Branch Code' and 'Maker' columns in Sheet1")

    esc = esc[[col_bc, col_maker]].copy()
    esc[col_bc] = esc[col_bc].astype(str).str.strip()
    esc[col_maker] = esc[col_maker].astype(str).str.strip()

    esc = esc[(esc[col_bc] != "") & (esc[col_bc].str.lower() != "nan")]
    return dict(zip(esc[col_bc], esc[col_maker]))


def build_required_format(df, escalation_map):
    col_zone = find_col(df, "zone")
    col_region = find_col(df, "Region")
    col_branch_code = find_col(df, "Branch Code")
    col_branch = find_col(df, "BRANCH_NAME")

    col_loan = find_col(df, "LOAN_ID")
    col_cust = find_col(df, "CUST_ID")
    col_date = find_col(df, "COLLECTION_DATE")

    col_amount = find_col(df, "Amount")
    col_principal = find_col(df, "PRINCIPAL_COLLECTED")
    col_interest = find_col(df, "INTEREST_COLLECTED")

    col_excess = find_col(df, "Excess Amount") or "Excess Amount"

    missing = []
    for nm, col in [
        ("zone", col_zone),
        ("Region", col_region),
        ("Branch Code", col_branch_code),
        ("BRANCH_NAME", col_branch),
        ("LOAN_ID", col_loan),
        ("CUST_ID", col_cust),
        ("COLLECTION_DATE", col_date),
        ("Amount", col_amount),
        ("PRINCIPAL_COLLECTED", col_principal),
        ("INTEREST_COLLECTED", col_interest),
    ]:
        if col is None:
            missing.append(nm)
    if missing:
        raise ValueError(f"Missing required columns for final format: {missing}")

    branch_code_series = df[col_branch_code].astype(str).str.strip()
    executive_series = branch_code_series.map(escalation_map).fillna("")

    out = pd.DataFrame({
        "Zone": df[col_zone],
        "Region": df[col_region],
        "Branch Code": df[col_branch_code],
        "Branch": df[col_branch],
        "Executive": executive_series,
        "Loan_Id": df[col_loan],
        "Cust_Id": df[col_cust],
        "Collection_Date": df[col_date],
        "Amount": df[col_amount],
        "Principal_Collected": df[col_principal],
        "Interest_Collected": df[col_interest],
        "Excess Amount": df[col_excess],
    })

    return out


def process_file(file_path: str, escalation_map, write_back=True):
    df, sheet_used = read_sheet_with_status(file_path)

    col_status = find_col(df, "Status")
    col_amount = find_col(df, "Amount")
    col_principal = find_col(df, "PRINCIPAL_COLLECTED")
    col_interest = find_col(df, "INTEREST_COLLECTED")
    col_lpc = find_col(df, "Late_Payment_charges")

    missing = []
    for name, col in [
        ("Status", col_status),
        ("Amount", col_amount),
        ("PRINCIPAL_COLLECTED", col_principal),
        ("INTEREST_COLLECTED", col_interest),
        ("Late_Payment_charges", col_lpc),
    ]:
        if col is None:
            missing.append(name)
    if missing:
        raise ValueError(f"Missing column(s) in {os.path.basename(file_path)}: {missing}")

    df2 = df[df[col_status].astype(str).str.strip().str.lower() == "processed"].copy()

    df2[col_amount] = to_num(df2[col_amount])
    df2[col_principal] = to_num(df2[col_principal])
    df2[col_interest] = to_num(df2[col_interest])
    df2[col_lpc] = to_num(df2[col_lpc])

    df2["Excess Amount"] = df2[col_amount] - (df2[col_principal] + df2[col_interest])

    df2 = df2[df2["Excess Amount"] >= 1].copy()
    df2 = df2[df2[col_lpc] <= 0].copy()

    # Optional: Write back sheet into uploaded file
    if write_back:
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df2.to_excel(writer, sheet_name=OUTPUT_SHEET_NAME, index=False)

    return build_required_format(df2, escalation_map)


def format_consolidated_excel(xlsx_path: str, sheet_name: str):
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]

    header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    max_col = ws.max_column

    ws.freeze_panes = "A2"

    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    data_alignment = Alignment(vertical="center")
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = data_alignment

    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        max_len = 0
        for r in range(1, max_row + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).strip() != "":
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    wb.save(xlsx_path)
    wb.close()


def run_excess_amount_from_streamlit(il_path: str, jlg_path: str, escalation_path: str, output_dir: str):
    """
    Streamlit-safe runner.
    - il_path: temp saved uploaded Repayment Summary IL.xlsx
    - jlg_path: temp saved uploaded Repayment Summary JLG.xlsx
    - escalation_path: temp saved uploaded Escalation.xlsx
    - output_dir: temp folder to write consolidated output
    Returns: consolidated_xlsx_path
    """
    escalation_map = load_escalation_map(escalation_path)

    all_data = []
    for f in [il_path, jlg_path]:
        req_df = process_file(f, escalation_map, write_back=True)
        all_data.append(req_df)

    consolidated = pd.concat(all_data, ignore_index=True)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    consolidated_file = os.path.join(output_dir, f"Excess Amount Received - Consolidated_{stamp}.xlsx")

    with pd.ExcelWriter(consolidated_file, engine="openpyxl", mode="w") as writer:
        consolidated.to_excel(writer, sheet_name=CONSOLIDATED_SHEET, index=False)

    format_consolidated_excel(consolidated_file, CONSOLIDATED_SHEET)

    return consolidated_file


import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

TOL = 0.000001


def to_num(x):
    try:
        if x is None or x == "":
            return 0.0
        return float(x)
    except:
        return 0.0


def norm(s):
    return str(s).strip().lower() if s is not None else ""


def is_branch_id(val):
    s = str(val).strip() if val is not None else ""
    return bool(re.match(r"^B\d+", s, re.IGNORECASE))


def format_difference_sheet(ws):
    header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    bold_font = Font(bold=True)

    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = header_fill

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    max_col = ws.max_column

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).border = border

    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        max_len = 0
        for r in range(1, max_row + 1):
            v = ws.cell(r, c).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

    for r in range(1, max_row + 1):
        ws.row_dimensions[r].height = 20

    ws.freeze_panes = "A2"


def create_consolidated_sheet_cpp_payable(cpp_payable_file_path):
    """
    Creates/updates 'Consolidated' sheet in CPP Payable file:
      - Column A = 'Closing Balance'
      - Column B starts with 'B' (Branch ID)
    Copies columns A..I as VALUES (no xlwings, works on Streamlit Cloud)
    """
    CONSOL_SHEET = "Consolidated"
    COL_REGION = "A"
    COL_BRANCH = "B"
    COPY_COLS = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]

    wb = load_workbook(cpp_payable_file_path, data_only=True)

    # remove old consolidated if exists
    if CONSOL_SHEET in wb.sheetnames:
        del wb[CONSOL_SHEET]

    ws_out = wb.create_sheet(CONSOL_SHEET, 0)
    ws_out.append(["Source Sheet"] + COPY_COLS)

    for sh in wb.sheetnames:
        if sh == CONSOL_SHEET:
            continue

        ws = wb[sh]

        for r in range(1, ws.max_row + 1):
            region_val = ws[f"{COL_REGION}{r}"].value
            branch_val = ws[f"{COL_BRANCH}{r}"].value

            if norm(region_val) == "closing balance" and is_branch_id(branch_val):
                row_data = [sh]
                for col in COPY_COLS:
                    row_data.append(ws[f"{col}{r}"].value)
                ws_out.append(row_data)

    wb.save(cpp_payable_file_path)


def run_cpp_payable_vs_cpp_ledger_difference(cpp_payable_file_path, cpp_ledger_file_path):
    CONSOL_SHEET = "Consolidated"
    LEDGER_SHEET = "Ledger"
    ESC_SHEET = "Escalation"
    DIFF_SHEET = "Difference"

    # Step 0: Create/Update Consolidated in CPP Payable file
    create_consolidated_sheet_cpp_payable(cpp_payable_file_path)

    # 1) Read CPP Payable Consolidated
    df_payable = pd.read_excel(cpp_payable_file_path, sheet_name=CONSOL_SHEET, engine="openpyxl")
    branch_col = "B"
    balance_col = "I"

    df_payable[branch_col] = df_payable[branch_col].astype(str).str.strip()
    df_payable[balance_col] = pd.to_numeric(df_payable[balance_col], errors="coerce").fillna(0)
    payable_balance_by_branch = df_payable.set_index(branch_col)[balance_col].to_dict()

    # 2) Open CPP Ledger workbook
    wb = load_workbook(cpp_ledger_file_path)
    if LEDGER_SHEET not in wb.sheetnames:
        raise Exception(f"'{LEDGER_SHEET}' sheet not found in CPP Ledger file.")
    if ESC_SHEET not in wb.sheetnames:
        raise Exception(f"'{ESC_SHEET}' sheet not found in CPP Ledger file.")

    ws_ledger = wb[LEDGER_SHEET]
    ws_esc = wb[ESC_SHEET]

    # 3) Escalation mapping Branch Code -> Maker (Executive Name)
    esc_headers = {}
    for c in range(1, ws_esc.max_column + 1):
        h = ws_esc.cell(1, c).value
        if h:
            esc_headers[norm(h)] = c

    col_branch_code = esc_headers.get("branch code")
    col_maker = esc_headers.get("maker")
    if not col_branch_code or not col_maker:
        raise Exception("Escalation sheet must have headers: 'Branch Code' and 'Maker' in row 1")

    exec_by_branch = {}
    for r in range(2, ws_esc.max_row + 1):
        bc = ws_esc.cell(r, col_branch_code).value
        mk = ws_esc.cell(r, col_maker).value
        if bc is None:
            continue
        exec_by_branch[str(bc).strip()] = "" if mk is None else str(mk).strip()

    # 4) Ledger columns
    ledger_headers = {}
    for c in range(1, ws_ledger.max_column + 1):
        h = ws_ledger.cell(1, c).value
        if h:
            ledger_headers[norm(h)] = c

    sub_acc_col = ledger_headers.get("sub account")
    debit_col = ledger_headers.get("closing debit")
    credit_col = ledger_headers.get("closing credit")
    if not (sub_acc_col and debit_col and credit_col):
        raise Exception("Ledger sheet must have headers: 'Sub Account', 'Closing Debit', 'Closing Credit' in row 1")

    # 5) Create/Replace Difference sheet
    if DIFF_SHEET in wb.sheetnames:
        del wb[DIFF_SHEET]
    ws_diff = wb.create_sheet(DIFF_SHEET)

    ws_diff.append([
        "Executive Name",
        "Branch Code",
        "Ledger Balance (Cl.Credit-Cl.Debit)",
        "CPP Payable Sheet Balance",
        "Difference"
    ])

    # 6) Only difference rows
    for r in range(2, ws_ledger.max_row + 1):
        sub_acc = ws_ledger.cell(r, sub_acc_col).value
        if sub_acc is None or str(sub_acc).strip() == "":
            continue

        key = str(sub_acc).strip()

        closing_credit = to_num(ws_ledger.cell(r, credit_col).value)
        closing_debit = to_num(ws_ledger.cell(r, debit_col).value)
        final_balance = closing_credit - closing_debit

        payable_balance = to_num(payable_balance_by_branch.get(key, 0))
        diff = final_balance - payable_balance

        if abs(diff) > TOL:
            exec_name = exec_by_branch.get(key, "")
            ws_diff.append([exec_name, key, final_balance, payable_balance, diff])

    format_difference_sheet(ws_diff)
    wb.save(cpp_ledger_file_path)

import os
import re
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# =========================
# LEDGER MAPPING RULES
# =========================
LEDGER_REMOVE_ACCOUNTS = {
    "roi net - c",
    "ease buzz-c",
    "cms deposit a/c"
}

ACCOUNT_RENAME_MAP = {
    "spice money - c": "SPICE",
    "airtel payments bank - c": "AIRTEL",
    "fino payments bank - c": "FINO",
    "fingpay - c": "FINGPAY",
    "axis bbps - c": "BBPS",
    "airpay-c": "AIRPAY",
    "idfc cms": "IDFC",
    "twinline  - c": "TWINLINE",
    "twinline - c": "TWINLINE",
}

# =========================
# HELPERS
# =========================
def safe_sheet_name_from_path(path: str) -> str:
    name = os.path.splitext(os.path.basename(path))[0]
    name = re.sub(r"[\[\]\:\*\?\/\\]", " ", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name[:31] if len(name) > 31 else name

def normalize_branch_code_generic(x) -> str:
    if x is None:
        return ""
    s = str(x).strip().upper()
    if not s:
        return ""

    if "MMF" in s:
        digits = re.sub(r"\D", "", s)
        digits = digits[-4:] if len(digits) >= 4 else digits.zfill(4)
        return "B" + digits.zfill(4)

    if "BRANCH" in s and "-" in s:
        part = s.split("-")[-1]
        digits = re.sub(r"\D", "", part)
        return "B" + digits.zfill(4)

    m = re.search(r"\bB\d{4}\b", s)
    if m:
        return m.group(0)

    digits = re.sub(r"\D", "", s)
    if not digits:
        return ""
    return "B" + digits.zfill(4)

def normalize_ledger_code_strict(x) -> str:
    if x is None:
        return ""
    s = str(x).strip().upper()
    if not s:
        return ""
    if s.startswith("HO"):
        return ""
    if re.fullmatch(r"B\d{4}", s):
        return s
    if re.fullmatch(r"\d{1,4}", s):
        return "B" + s.zfill(4)
    digits = re.sub(r"\D", "", s)
    if digits and len(digits) <= 4:
        return "B" + digits.zfill(4)
    return ""

def apply_formatting(ws, last_row: int, last_col: int):
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    header_font = Font(bold=True, color="1F4E79")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(1, last_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for r in range(2, last_row + 1):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = center

    ws.freeze_panes = "A2"

    for c in range(1, last_col + 1):
        col_letter = get_column_letter(c)
        max_len = 0
        for r in range(1, last_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)

def list_statement_files(folder: str):
    # Streamlit Cloud safe: prefer .xlsx/.csv. .xls often fails due to xlrd not installed.
    exts = (".xlsx", ".csv", ".xls")
    files = []
    for f in os.listdir(folder):
        if f.lower().endswith(exts):
            files.append(os.path.join(folder, f))
    files.sort()
    return files

def _clean_colname(c) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(c).strip().lower())

def find_col_fuzzy(df: pd.DataFrame, expected: str):
    exp = _clean_colname(expected)
    for c in df.columns:
        if _clean_colname(c) == exp:
            return c
    for c in df.columns:
        if exp in _clean_colname(c):
            return c
    return None

def build_branch_amount_map(df: pd.DataFrame, branch_col: str, amount_col: str) -> dict:
    work = df[[branch_col, amount_col]].copy()
    work[branch_col] = work[branch_col].apply(normalize_branch_code_generic)
    work[amount_col] = pd.to_numeric(work[amount_col], errors="coerce").fillna(0)
    grp = work.groupby(branch_col, dropna=False)[amount_col].sum()
    return {str(k): float(v) for k, v in grp.items() if str(k).strip()}

def read_statement(path: str) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()

    if ext == ".csv":
        return pd.read_csv(path)

    if ext == ".xlsx":
        return pd.read_excel(path, engine="openpyxl")

    if ext == ".xls":
        # Cloud-safe: xlrd may not exist. So fail with a clear message.
        raise ValueError(f".xls file not supported on Streamlit Cloud: {os.path.basename(path)}. Please convert to .xlsx or .csv")

    raise ValueError(f"Unsupported file type: {path}")

def _filter_success_generic(df: pd.DataFrame, status_col: str) -> pd.DataFrame:
    s = df[status_col].astype(str).str.strip().str.lower()
    ok = s.str.contains("success", na=False) | s.isin(["successful", "success", "transaction successful"])
    bad = s.str.contains("fail", na=False) | s.str.contains("revers", na=False)
    return df[ok & (~bad)].copy()

# =========================
# STATEMENT EXTRACTION
# =========================
def extract_for_sheet(sheet_name: str, file_path: str):
    try:
        df = read_statement(file_path)
        s = sheet_name.strip().lower()

        if s == "airpay":
            status_col = find_col_fuzzy(df, "Transaction Status") or find_col_fuzzy(df, "Status")
            branch_col = find_col_fuzzy(df, "Branch Code")
            amount_col = find_col_fuzzy(df, "Amount")
            if not status_col or not branch_col or not amount_col:
                return {}, 0.0, "AIRPAY: columns missing"
            df[status_col] = df[status_col].astype(str).str.strip().str.lower()
            df_f = df[df[status_col] != "failed"].copy()
            total = pd.to_numeric(df_f[amount_col], errors="coerce").fillna(0).sum()
            return build_branch_amount_map(df_f, branch_col, amount_col), float(total), ""

        if s == "airtel":
            amount_col = find_col_fuzzy(df, "ORIG_AMNT")
            branch_col = find_col_fuzzy(df, "Additional Information 2")
            if not branch_col or not amount_col:
                return {}, 0.0, "AIRTEL: columns missing"
            total = pd.to_numeric(df[amount_col], errors="coerce").fillna(0).sum()
            return build_branch_amount_map(df, branch_col, amount_col), float(total), ""

        if s == "bbps":
            status_col = find_col_fuzzy(df, "Payment Status")
            branch_col = find_col_fuzzy(df, "B CODE")
            amount_col = find_col_fuzzy(df, "Bill Amount") or find_col_fuzzy(df, "Transaction Amount")
            if not status_col or not branch_col or not amount_col:
                return {}, 0.0, "BBPS: columns missing"
            df[status_col] = df[status_col].astype(str).str.strip().str.lower()
            df_f = df[df[status_col] == "successful"].copy()
            total = pd.to_numeric(df_f[amount_col], errors="coerce").fillna(0).sum()
            return build_branch_amount_map(df_f, branch_col, amount_col), float(total), ""

        if s == "fingpay":
            status_col = find_col_fuzzy(df, "Status Message")
            amount_col = find_col_fuzzy(df, "Drop Amount")
            branch_col = find_col_fuzzy(df, "Branch Code")
            if not status_col or not branch_col or not amount_col:
                return {}, 0.0, "FINGPAY: columns missing"
            df[status_col] = df[status_col].astype(str).str.strip().str.lower()
            df_f = df[df[status_col].str.contains("success", na=False)].copy()
            total = pd.to_numeric(df_f[amount_col], errors="coerce").fillna(0).sum()
            return build_branch_amount_map(df_f, branch_col, amount_col), float(total), ""

        if s == "spice":
            status_col = find_col_fuzzy(df, "Status")
            amount_col = find_col_fuzzy(df, "Amount")
            branch_col = find_col_fuzzy(df, "Branch ID")
            if not status_col or not branch_col or not amount_col:
                return {}, 0.0, "SPICE: columns missing"
            df[status_col] = df[status_col].astype(str).str.strip().str.upper()
            df_f = df[df[status_col] == "SUCCESS"].copy()
            total = pd.to_numeric(df_f[amount_col], errors="coerce").fillna(0).sum()
            return build_branch_amount_map(df_f, branch_col, amount_col), float(total), ""
            if s == "fino":
    status_col = find_col_fuzzy(df, "Status")
    amount_col = find_col_fuzzy(df, "AMOUNT")
    branch_col = find_col_fuzzy(df, "Branch ID")

    if not status_col or not branch_col or not amount_col:
        return {}, 0.0, f"FINO: columns missing. Found: {list(df.columns)}"

    df[status_col] = df[status_col].astype(str).str.strip().str.lower()
    df_f = df[
        df[status_col].str.contains("successful", na=False)
        | df[status_col].str.contains("success", na=False)
    ].copy()

    total = pd.to_numeric(df_f[amount_col], errors="coerce").fillna(0).sum()
    return build_branch_amount_map(df_f, branch_col, amount_col), float(total), ""
  

        
        if s == "twinline":
            branch_candidates = ["Branch Code", "Branch", "Branch ID", "B CODE", "Code", "Additional Information 2"]
            amount_candidates = ["Amount", "AMOUNT", "Transaction Amount", "Bill Amount", "ORIG_AMNT", "Drop Amount", "Net", "Net Amount"]
            status_candidates = ["Status", "Transaction Status", "Payment Status", "Status Message"]

            branch_col = None
            for x in branch_candidates:
                branch_col = find_col_fuzzy(df, x)
                if branch_col:
                    break

            amount_col = None
            for x in amount_candidates:
                amount_col = find_col_fuzzy(df, x)
                if amount_col:
                    break

            status_col = None
            for x in status_candidates:
                status_col = find_col_fuzzy(df, x)
                if status_col:
                    break

            if not branch_col or not amount_col:
                return {}, 0.0, f"TWINLINE: branch/amount columns not found. Found: {list(df.columns)}"

            df_f = df.copy()
            if status_col:
                df_f = _filter_success_generic(df_f, status_col)

            total = pd.to_numeric(df_f[amount_col], errors="coerce").fillna(0).sum()
            return build_branch_amount_map(df_f, branch_col, amount_col), float(total), ""

        return {}, 0.0, "NOT CONFIGURED"

    except Exception as e:
        return {}, 0.0, str(e)

# =========================
# LEDGER: Branch mapping
# =========================
def normalize_branch_from_branchid(x) -> str:
    if x is None:
        return ""
    s = str(x).strip().upper()
    m = re.search(r"\bB\d{4}\b", s)
    return m.group(0) if m else ""

def extract_bcode_from_description(x) -> str:
    if x is None:
        return ""
    s = str(x).upper()
    m = re.search(r"\bB\d{4}\b", s)
    return m.group(0) if m else ""

def ledger_bcode_rowwise(row, branchid_col=None, desc_col=None, code_col=None) -> str:
    if branchid_col and branchid_col in row and pd.notna(row[branchid_col]):
        b = normalize_branch_from_branchid(row[branchid_col])
        if b:
            return b
    if desc_col and desc_col in row and pd.notna(row[desc_col]):
        b = extract_bcode_from_description(row[desc_col])
        if b:
            return b
    if code_col and code_col in row and pd.notna(row[code_col]):
        b = normalize_ledger_code_strict(row[code_col])
        if b:
            return b
    return ""

def build_ledger_maps(cms_ledger_path: str):
    df = pd.read_excel(cms_ledger_path, engine="openpyxl")

    acc_col = find_col_fuzzy(df, "Account Name")
    code_col = find_col_fuzzy(df, "code")
    debit_col = find_col_fuzzy(df, "Debit")
    credit_col = find_col_fuzzy(df, "Credit")
    desc_col = find_col_fuzzy(df, "Description")
    branchid_col = find_col_fuzzy(df, "Branch ID")

    if not acc_col or not code_col or not debit_col or not credit_col:
        raise ValueError("CMS Ledger: required columns not found (Account Name, code, Debit, Credit).")

    df[acc_col] = df[acc_col].astype(str).str.strip()
    df["_acc_l"] = df[acc_col].str.lower()

    df = df[~df["_acc_l"].isin(LEDGER_REMOVE_ACCOUNTS)].copy()

    df["_sheet"] = df["_acc_l"].map(ACCOUNT_RENAME_MAP).fillna(df[acc_col])
    df["_sheet"] = df["_sheet"].astype(str).str.strip().str.upper()

    df[debit_col] = pd.to_numeric(df[debit_col], errors="coerce").fillna(0)
    df[credit_col] = pd.to_numeric(df[credit_col], errors="coerce").fillna(0)
    df["Net"] = df[debit_col] - df[credit_col]

    df["_bcode"] = df.apply(
        lambda r: ledger_bcode_rowwise(r, branchid_col=branchid_col, desc_col=desc_col, code_col=code_col),
        axis=1
    )
    df = df[df["_bcode"] != ""].copy()

    ledger_maps = {}
    ledger_totals = {}

    for sheet, g in df.groupby("_sheet"):
        m = g.groupby("_bcode")["Net"].sum().to_dict()
        ledger_maps[sheet] = {str(k): float(v) for k, v in m.items()}
        ledger_totals[sheet] = float(sum(m.values()))

    return ledger_maps, ledger_totals

# =========================
# MAIN RUNNER
# =========================
def run_all_statements_with_ledger(format_file: str, statements_folder: str, ledger_file: str, output_file: str):
    wb_format = load_workbook(format_file)
    template_ws = wb_format.worksheets[0]
    template_data = list(template_ws.values)

    if not template_data or len(template_data) < 2:
        raise ValueError("Format file first sheet is empty or has no data.")

    static_rows = []
    for row in template_data[1:]:
        r = list(row[:4])
        while len(r) < 4:
            r.append("")
        static_rows.append(r)

    ledger_maps, ledger_totals = build_ledger_maps(ledger_file)

    statement_files = list_statement_files(statements_folder)
    if not statement_files:
        raise ValueError(f"No statement files found in folder: {statements_folder}")

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    ws_sum = wb_out.create_sheet("Summary", 0)
    ws_sum.append(["Sheet", "File", "Statement Total", "Pasted Statement",
                   "Ledger Total", "Pasted Ledger", "Difference", "Status", "Remarks/Error"])

    headers = ["Region", "Branch Code", "Branch Name", "Maker", "Statement", "Ledger", "Difference"]
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")

    summary_rows = []

    for fp in statement_files:
        sh_name = safe_sheet_name_from_path(fp).strip().upper()

        ws = wb_out.create_sheet(sh_name[:31])
        ws.append(headers)

        stmt_map, stmt_total, stmt_err = extract_for_sheet(sh_name.lower(), fp)

        led_map = ledger_maps.get(sh_name, {})
        led_total = ledger_totals.get(sh_name, 0.0)

        pasted_stmt_total = 0.0
        pasted_led_total = 0.0

        for r in static_rows:
            region, bcode, bname, maker = r
            bcode_norm = normalize_branch_code_generic(bcode)

            stmt_val = float(stmt_map.get(bcode_norm, 0)) if stmt_map else 0.0
            led_val = float(led_map.get(bcode_norm, 0)) if led_map else 0.0
            diff = stmt_val - led_val

            pasted_stmt_total += stmt_val
            pasted_led_total += led_val

            ws.append([region, bcode_norm, bname, maker, stmt_val, led_val, diff])

        apply_formatting(ws, ws.max_row, ws.max_column)

        if not stmt_map:
            status = "STATEMENT NOT EXTRACTED"
            remarks = stmt_err
        else:
            stmt_ok = abs(stmt_total - pasted_stmt_total) <= 0.01
            status = "MATCHED" if stmt_ok else "STATEMENT TOTAL NOT MATCHED"
            remarks = "" if stmt_ok else f"Statement total mismatch: file={stmt_total:.2f}, pasted={pasted_stmt_total:.2f}"

            if led_map:
                led_ok = abs(led_total - pasted_led_total) <= 0.01
                if not led_ok:
                    status += " | LEDGER TOTAL NOT MATCHED"
                    remarks = (remarks + " | " if remarks else "") + f"Ledger total mismatch: file={led_total:.2f}, pasted={pasted_led_total:.2f}"
            else:
                status += " | LEDGER NOT FOUND"
                remarks = (remarks + " | " if remarks else "") + "Ledger mapping not found for this sheet (Account Name not mapped)."

        summary_rows.append([
            sh_name,
            os.path.basename(fp),
            round(stmt_total, 2),
            round(pasted_stmt_total, 2),
            round(led_total, 2),
            round(pasted_led_total, 2),
            round(pasted_stmt_total - pasted_led_total, 2),
            status,
            remarks
        ])

    for row in summary_rows:
        ws_sum.append(row)

    apply_formatting(ws_sum, ws_sum.max_row, ws_sum.max_column)

    for r in range(2, ws_sum.max_row + 1):
        st = str(ws_sum.cell(r, 8).value).upper()
        if "MATCHED" in st and "NOT" not in st and "MISMATCH" not in st:
            fill = green_fill
        elif "NOT" in st or "MISMATCH" in st:
            fill = red_fill
        else:
            fill = yellow_fill
        for c in range(1, 10):
            ws_sum.cell(r, c).fill = fill

    wb_out.save(output_file)
    return output_file


def run_cms_recon_streamlit(format_file: str, statements_folder: str, cms_ledger_file: str, output_file: str) -> str:
    return run_all_statements_with_ledger(format_file, statements_folder, cms_ledger_file, output_file)


# ============================================================
# PAGE UI
# ============================================================
st.set_page_config(page_title="Collection Efficiency Automation", page_icon="📊", layout="wide")
st.title("📊 Reporting Team- Reports Automisation")
st.markdown("""
<style>
/* --- TAB BUTTON STYLE (Compact + Bright Orange) --- */

/* Tab container */
.stTabs [data-baseweb="tab-list"]{
    gap: 4px;
    border-bottom: none !important;
    padding-bottom: 4px;
    flex-wrap: nowrap;
    overflow-x: auto;
}

/* Each tab */
.stTabs [data-baseweb="tab"]{
    background: #2b2b2b !important;          /* dark grey */
    border: 1px solid #ffffff !important;    /* white outline */
    border-radius: 4px !important;
    padding: 3px 6px !important;             /* smaller width */
    color: #ffffff !important;
    min-height: 26px !important;
}

/* Tab text */
.stTabs [data-baseweb="tab"] p{
    font-size: 11px !important;
    font-weight: 600 !important;
    margin: 0 !important;
    white-space: nowrap;
}

/* Active tab – Bright Orange */
.stTabs [aria-selected="true"]{
    background: #ff6a00 !important;          /* sharp bright orange */
    border: 1px solid #ff6a00 !important;
    color: #ffffff !important;
}

/* Remove default underline */
.stTabs [data-baseweb="tab-highlight"],
.stTabs [data-baseweb="tab-border"]{
    display: none !important;
}

/* ---------------- File Uploader Button Styling ---------------- */

div[data-testid="stFileUploader"] button{
    background-color: #ff6a00 !important;    /* same bright orange */
    color: white !important;
    border-radius: 6px !important;
    border: none !important;
    padding: 5px 12px !important;            /* compact button */
    font-weight: 600 !important;
}

div[data-testid="stFileUploader"] button:hover{
    background-color: #e65c00 !important;
}

/* ============================================================
   NEW: Make Upload Boxes Smaller (Height + Padding) everywhere
   ============================================================ */

/* Orange label line: "Upload ....xlsx" */
div[data-testid="stFileUploader"] label{
    color: #ff6a00 !important;               /* same as Browse button */
    font-weight: 700 !important;
    font-size: 13px !important;
}

/* Reduce uploader drop-zone height + padding */
div[data-testid="stFileUploader"] section{
    padding: 8px 10px !important;            /* smaller padding */
    min-height: 52px !important;             /* reduce height */
    border-radius: 10px !important;
}

/* Make inner content compact */
div[data-testid="stFileUploader"] section *{
    line-height: 1.15 !important;
}

/* "Drag and drop..." and helper text smaller */
div[data-testid="stFileUploader"] section small{
    font-size: 10px !important;
}

/* Uploaded file row compact */
div[data-testid="stFileUploader"] ul{
    margin-top: 4px !important;
}
div[data-testid="stFileUploader"] li{
    padding: 4px 8px !important;
    font-size: 11px !important;
    border-radius: 8px !important;
}

/* Optional: make browse button slightly smaller too */
div[data-testid="stFileUploader"] button{
    padding: 4px 10px !important;
    font-size: 13px !important;
}
/* ===== Force ultra-compact upload box ===== */

/* Outer wrapper */
div[data-testid="stFileUploader"]{
    margin-bottom: 6px !important;
}

/* Drop zone main area */
div[data-testid="stFileUploader"] section{
    min-height: 38px !important;     /* hard reduce height */
    padding: 6px 8px !important;     /* tight padding */
}

/* Icon + text row */
div[data-testid="stFileUploader"] section > div{
    gap: 6px !important;
}

/* Cloud icon size */
div[data-testid="stFileUploader"] svg{
    width: 20px !important;
    height: 20px !important;
}

/* Main text: "Drag and drop file here" */
div[data-testid="stFileUploader"] section strong,
div[data-testid="stFileUploader"] section span{
    font-size: 12px !important;
    line-height: 1.1 !important;
}

/* Helper text: "Limit 200MB..." */
div[data-testid="stFileUploader"] section small{
    font-size: 9px !important;
    margin-top: 0 !important;
    line-height: 1 !important;
}

/* Uploaded file row */
div[data-testid="stFileUploader"] li{
    padding: 3px 6px !important;
    font-size: 10px !important;
}
/* ===== HARD REDUCE uploader box HEIGHT (not only text) ===== */

/* Main drop area */
div[data-testid="stFileUploader"] section{
    padding: 4px 8px !important;
    min-height: 34px !important;
    height: 44px !important;          /* force actual height */
    display: flex !important;
    align-items: center !important;
}

/* Inner wrapper */
div[data-testid="stFileUploader"] section > div{
    padding: 0 !important;
    margin: 0 !important;
    width: 100% !important;
}

/* The clickable dropzone block */
div[data-testid="stFileUploader"] section div[role="button"]{
    padding: 0 !important;
    margin: 0 !important;
    min-height: 34px !important;
    height: 44px !important;          /* force height */
    display: flex !important;
    align-items: center !important;
}

/* Reduce left icon spacing */
div[data-testid="stFileUploader"] section svg{
    width: 18px !important;
    height: 18px !important;
    margin-right: 6px !important;
}
/* ===== Compact Browse Files button container ===== */

/* Reduce the height of the button wrapper */
div[data-testid="stFileUploader"] section div[role="button"] > div:last-child{
    padding: 0 !important;
    margin: 0 !important;
    display: flex !important;
    align-items: center !important;
}

/* Make the Browse button slimmer */
div[data-testid="stFileUploader"] button{
    padding: 3px 10px !important;   /* smaller height */
    font-size: 12px !important;
    line-height: 1.1 !important;
    border-radius: 5px !important;
}
/* ========= Report Dropdown Styling ========= */
div[data-testid="stSelectbox"] label{
    color: #ff6a00 !important;
    font-weight: 800 !important;
    font-size: 14px !important;
}

div[data-testid="stSelectbox"] div[role="combobox"]{
    background: #2b2b2b !important;
    border: 1px solid #ffffff !important;
    border-radius: 8px !important;
    padding: 6px 10px !important;
    min-height: 42px !important;
    color: #ffffff !important;
}

div[data-testid="stSelectbox"] div[role="combobox"]:focus-within{
    border: 1px solid #ff6a00 !important;
    box-shadow: 0 0 0 2px rgba(255,106,0,0.35) !important;
}
/* ===== Selectbox (Report Dropdown) – Match Browse Button Color ===== */

/* Main selectbox container */
div[data-baseweb="select"] > div {
    background-color: #ff6a00 !important;
    border: 1px solid #ff6a00 !important;
    color: white !important;
}

/* Text inside selectbox */
div[data-baseweb="select"] span {
    color: white !important;
    font-weight: 600 !important;
}

/* Dropdown arrow */
div[data-baseweb="select"] svg {
    fill: white !important;
}

/* Dropdown menu items */
ul[role="listbox"] {
    background-color: #1f1f1f !important;
}

/* Hovered option */
ul[role="listbox"] li:hover {
    background-color: #ff6a00 !important;
    color: white !important;
}

/* Selected option */
ul[role="listbox"] li[aria-selected="true"] {
    background-color: #ff6a00 !important;
    color: white !important;
}
/* Reduce width of report dropdown (search bar) */
div[data-testid="stSelectbox"]{
    max-width: 50% !important;   /* half width */
}

</style>
""", unsafe_allow_html=True)

report_options = [
    "1) Collection Efficiency Report",
    "2) Arrear Advance Report",
    "3) LPC Report",
    "4) Vault Cash Report",
    "5) Pending Collection Entry",
    "6) Demand Verification Report",
    "7) MAP Ledger Difference",
    "8) Excess Amount",
    "9) CPP Payable",
    "10) CMS and UPI Recon"
]

selected_report = st.selectbox(
    "🔎 Select Report (type to search)",
    report_options,
    index=0,
    key="report_selector"
)


# ============================================================
# HELPERS
# ============================================================
def save_uploaded_file(uploaded, path):
    with open(path, "wb") as f:
        f.write(uploaded.getbuffer())


def zip_folder(folder_path, zip_path):
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, _, files in os.walk(folder_path):
            for file in files:
                full_path = os.path.join(root, file)
                rel_path = os.path.relpath(full_path, folder_path)
                zf.write(full_path, rel_path)


def clean_name_for_folder(text):
    if text is None:
        return "BLANK"
    s = str(text).strip()
    if s == "" or s.lower() == "nan":
        return "BLANK"
    s = re.sub(r'[\\/:*?"<>|]', "", s)
    return s[:80]


def safe_sheet_title(name: str) -> str:
    name = re.sub(r'[:\\/?*\[\]]', "_", str(name))
    return name[:31]


def to_numeric_clean(series: pd.Series) -> pd.Series:
    return pd.to_numeric(
        series.astype(str)
              .str.replace(",", "", regex=False)
              .str.replace("%", "", regex=False),
        errors="coerce"
    )


def is_percent_column(col_name: str, series_numeric: pd.Series) -> bool:
    name = str(col_name).lower()
    if "%" in name or "efficiency" in name or "pct" in name or "percent" in name or "(y/x)" in name:
        return True
    s = series_numeric.dropna()
    if s.empty:
        return False
    mn, mx = s.min(), s.max()
    if mn >= -1 and mx <= 1.5 and (s % 1 != 0).any():
        return True
    return False


def recompute_grand_total(df: pd.DataFrame, hub_col: str) -> pd.DataFrame:
    if df.empty:
        return df

    first_col = df.columns[0]
    gt_mask = df[first_col].astype(str).str.strip().str.lower().eq("grand total")
    if not gt_mask.any():
        return df

    gt_row = df[gt_mask].iloc[:1].copy()
    data_rows = df[~gt_mask].copy()

    for col in df.columns:
        if col in (first_col, hub_col):
            continue
        num = to_numeric_clean(data_rows[col])
        if num.notna().sum() == 0:
            continue
        if is_percent_column(col, num):
            gt_row.iloc[0, df.columns.get_loc(col)] = float(num.mean())
        else:
            gt_row.iloc[0, df.columns.get_loc(col)] = float(num.sum())

    return pd.concat([gt_row, data_rows], ignore_index=True)


def build_format_map(df: pd.DataFrame, hub_col: str) -> dict:
    fmt_map = {}
    for idx, col in enumerate(df.columns, start=1):
        if col == hub_col:
            continue
        num = to_numeric_clean(df[col])
        if num.notna().sum() == 0:
            continue
        fmt_map[idx] = "0.00%" if is_percent_column(col, num) else "#,##0"
    return fmt_map


def write_df_fast(ws, df: pd.DataFrame, format_map: dict):
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    max_row = ws.max_row
    for col_idx, fmt in format_map.items():
        for r in range(2, max_row + 1):
            c = ws.cell(row=r, column=col_idx)
            if isinstance(c.value, (int, float)) and c.value is not None:
                c.number_format = fmt


def read_ce_sheets(path: str) -> dict:
    return pd.read_excel(path, sheet_name=None, engine="openpyxl")


def read_csv_with_fallback(path, compression=None):
    encodings = ["utf-8", "utf-8-sig", "cp1252", "latin1"]
    last_err = None
    for enc in encodings:
        try:
            return pd.read_csv(path, compression=compression, low_memory=False, encoding=enc)
        except Exception as e:
            last_err = e
    raise last_err


def read_arrear_any(path: str, original_name: str, use_csv: bool) -> pd.DataFrame:
    ext = os.path.splitext(original_name)[1].lower()

    if use_csv and ext == ".csv":
        return read_csv_with_fallback(path)

    if use_csv and ext == ".zip":
        return read_csv_with_fallback(path, compression="zip")

    return pd.read_excel(path, engine="openpyxl")


# ============================================================
# DVC CONSOLIDATE LOGIC (Reusable)
# ============================================================
def build_dvc_consolidate(jlg_path: str, il_path: str) -> pd.DataFrame:
    header_map_jlg = [
        'COMPANY_ID', 'ZoneName', 'region_name', 'unit_name', 'cluster_name',
        'state_id', 'branch_id', 'branch_name', 'CENTER_ID', 'center_name',
        'group_name', 'FirstNAme', 'LastName', 'facility_id', 'installment_no',
        'loan_id', 'cust_id', 'opening_principal_advance', 'opening_interest_advance',
        'overdue_principal_demand', 'overdue_interest_demand', 'current_principal_demand',
        'current_interest_demand', 'insurance_premium_due', 'total_demand',
        'overdue_principal_collected', 'overdue_interest_collected', 'current_principal_collected',
        'current_interest_collected', 'advance_principal_collected', 'advance_interest_collected',
        'insurance_premium_collected', 'total_collection', 'excess_amt_collected',
        'disbursement_date', 'loan_maturity_date', 'Last_Collection_Paid', 'collection_date',
        'repayment_frequency', 'Loan_tenure_Month', 'No_of_Inst_OD', 'Total_Inst_Amt_Paid',
        'No_of_Inst_Due', 'mobile_number', 'P2P', 'prod_category_id', 'product_id',
        'status', 'Installment_Date', 'ReportDate', 'PRECLOSURE_REASON',
        'od_days', 'Spouse_Name'
    ]

    jlg_df = pd.read_excel(jlg_path, engine="openpyxl", usecols=header_map_jlg)
    il_df = pd.read_excel(il_path, engine="openpyxl")

    il_df.rename(columns={
        'zone_name': 'ZoneName',
        'area_name': 'unit_name',
        'report_date': 'ReportDate'
    }, inplace=True)

    if 'branch_name' in il_df.columns:
        il_df['branch_id'] = il_df['branch_name'].astype(str).str[:5]
    else:
        il_df['branch_id'] = None

    il_aligned_df = pd.DataFrame({c: il_df[c] if c in il_df.columns else None for c in header_map_jlg})
    consolidated_df = pd.concat([jlg_df, il_aligned_df], ignore_index=True)

    consolidated_df.columns = consolidated_df.columns.str.strip().str.replace(' ', '_').str.lower()

    if 'branch_id' in consolidated_df.columns:
        cols = consolidated_df.columns.tolist()
        cols.insert(0, cols.pop(cols.index('branch_id')))
        consolidated_df = consolidated_df[cols]

    if 'zonename' in consolidated_df.columns:
        consolidated_df.rename(columns={'zonename': 'hub'}, inplace=True)

    if 'company_id' in consolidated_df.columns:
        consolidated_df.drop(columns=['company_id'], inplace=True)

    if 'status' in consolidated_df.columns:
        consolidated_df = consolidated_df[
            ~consolidated_df['status'].astype(str).str.lower().str.contains('death', na=False)
        ]

    num_cols = [
        'overdue_interest_collected', 'current_interest_collected',
        'overdue_principal_collected', 'current_principal_collected',
        'opening_interest_advance', 'opening_principal_advance',
        'current_interest_demand', 'current_principal_demand',
        'advance_interest_collected', 'advance_principal_collected',
        'total_collection'
    ]
    for c in num_cols:
        if c not in consolidated_df.columns:
            consolidated_df[c] = 0
        consolidated_df[c] = pd.to_numeric(
            consolidated_df[c].astype(str).str.replace(",", "", regex=False),
            errors="coerce"
        ).fillna(0)

    mask_i = consolidated_df['current_interest_collected'] < 0
    consolidated_df.loc[mask_i, 'overdue_interest_collected'] += consolidated_df.loc[mask_i, 'current_interest_collected']
    consolidated_df.loc[mask_i, 'current_interest_collected'] = 0

    mask_p = consolidated_df['current_principal_collected'] < 0
    consolidated_df.loc[mask_p, 'overdue_principal_collected'] += consolidated_df.loc[mask_p, 'current_principal_collected']
    consolidated_df.loc[mask_p, 'current_principal_collected'] = 0

    consolidated_df['advance_demand'] = consolidated_df['opening_interest_advance'] + consolidated_df['opening_principal_advance']
    consolidated_df['current_demand'] = consolidated_df['current_interest_demand'] + consolidated_df['current_principal_demand']
    consolidated_df['advance'] = consolidated_df['advance_interest_collected'] + consolidated_df['advance_principal_collected']
    consolidated_df['advance'] = consolidated_df[['advance', 'total_collection']].min(axis=1)

    consolidated_df['collection_exc_advance'] = consolidated_df['total_collection'] - consolidated_df['advance']
    consolidated_df['same_day_demand'] = (consolidated_df['current_demand'] - consolidated_df['advance_demand']).clip(lower=0)
    consolidated_df.loc[consolidated_df['current_demand'] == 0, 'same_day_demand'] = 0
    consolidated_df['same_day_repayment_received'] = consolidated_df[['same_day_demand', 'collection_exc_advance']].min(axis=1)
    consolidated_df['od_received'] = (consolidated_df['collection_exc_advance'] - consolidated_df['same_day_demand']).clip(lower=0)
    consolidated_df['same_day_arrear'] = (consolidated_df['same_day_demand'] - consolidated_df['same_day_repayment_received']).clip(lower=0)

    return consolidated_df


def build_arrear_advance_sheet(consolidated_df: pd.DataFrame) -> pd.DataFrame:
    df = consolidated_df.copy()

    for col in ["cust_id", "advance", "same_day_arrear", "same_day_demand"]:
        if col not in df.columns:
            df[col] = 0

    df["advance"] = pd.to_numeric(df["advance"], errors="coerce").fillna(0)
    df["same_day_arrear"] = pd.to_numeric(df["same_day_arrear"], errors="coerce").fillna(0)
    df["same_day_demand"] = pd.to_numeric(df["same_day_demand"], errors="coerce").fillna(0)

    df = df[df["same_day_demand"] != 0].copy()

    pivot = (
        df.groupby("cust_id", as_index=False)[["advance", "same_day_arrear"]]
          .sum()
    )

    pivot = pivot[(pivot["advance"] > 0) & (pivot["same_day_arrear"] > 0)].copy()
    pivot.sort_values(["same_day_arrear", "advance"], ascending=False, inplace=True)

    return pivot


# ============================================================
# LPC REPORT MODULE (NEW)
# ============================================================
def lpc_generate_report(in_path: str, out_path: str, sheet_name=0):
    # ---------- Read ----------
    df = pd.read_excel(in_path, sheet_name=sheet_name, engine="openpyxl")
    df.columns = df.columns.astype(str).str.strip()

    # ---------- Filters ----------
    if "Status" not in df.columns:
        raise ValueError("Column 'Status' not found in LPC file.")
    df["Status"] = df["Status"].astype(str).str.strip().str.upper()
    df = df[~df["Status"].isin(["DEATH", "DEATH APPROVED"])].copy()

    if "DISBURSEMENT_DATE" not in df.columns:
        raise ValueError("Column 'DISBURSEMENT_DATE' not found in LPC file.")
    cutoff = pd.Timestamp("2025-10-01")
    raw_date = df["DISBURSEMENT_DATE"]

    dt1 = pd.to_datetime(raw_date, errors="coerce", dayfirst=True)
    is_serial = pd.to_numeric(raw_date, errors="coerce").notna() & dt1.isna()
    serial_vals = pd.to_numeric(raw_date, errors="coerce")

    dt2 = dt1.copy()
    dt2.loc[is_serial] = pd.to_datetime(serial_vals.loc[is_serial], unit="D", origin="1899-12-30", errors="coerce")
    df["DISBURSEMENT_DATE"] = dt2
    df = df[df["DISBURSEMENT_DATE"] >= cutoff].copy()

    # ---------- Remarks ----------
    for col in ["Total_Arrear", "Total_OS", "Pending_LPC"]:
        if col not in df.columns:
            raise ValueError(f"Missing column for remarks: {col}")
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df["Remarks"] = np.select(
        [
            (df["Pending_LPC"] <= 0),
            (df["Total_Arrear"] > 0) & (df["Total_OS"] > 0) & (df["Pending_LPC"] > 0),
            (df["Total_Arrear"] == 0) & (df["Total_OS"] == 0) & (df["Pending_LPC"] > 0),
            (df["Total_Arrear"] == 0) & (df["Total_OS"] > 0) & (df["Pending_LPC"] > 0),
        ],
        [
            "LPC Fully Collected",
            "Arrear & LPC Pending",
            "No Arrear & Outstanding but LPC Pending",
            "No Arrear but LPC Pending",
        ],
        default=""
    )

    # ---------- Pivot base ----------
    required = ["ZONE_NAME", "CLUSTER_NAME", "LOAN_ID", "LPC_Collected", "Pending_LPC"]
    for c in required:
        if c not in df.columns:
            raise ValueError(f"Missing column: {c}")

    df["LPC_Collected"] = pd.to_numeric(df["LPC_Collected"], errors="coerce").fillna(0)
    df["Pending_LPC"] = pd.to_numeric(df["Pending_LPC"], errors="coerce").fillna(0)

    if "Total_LPC_Due" not in df.columns:
        df["Total_LPC_Due"] = df["LPC_Collected"] + df["Pending_LPC"]
    else:
        df["Total_LPC_Due"] = pd.to_numeric(df["Total_LPC_Due"], errors="coerce").fillna(0)

    remark_order = [
        "Arrear & LPC Pending",
        "LPC Fully Collected",
        "No Arrear & Outstanding but LPC Pending",
        "No Arrear but LPC Pending"
    ]

    keys = df.groupby(["ZONE_NAME", "CLUSTER_NAME"], as_index=False).size()[["ZONE_NAME", "CLUSTER_NAME"]]
    summary_df = keys.copy()

    def block(rmk):
        d = df[df["Remarks"] == rmk]
        if d.empty:
            return pd.DataFrame(columns=["ZONE_NAME", "CLUSTER_NAME", "LPC Due", "LPC Collected", "LPC pending", "No. of loans"])
        return d.groupby(["ZONE_NAME", "CLUSTER_NAME"], as_index=False).agg(
            **{
                "LPC Due": ("Total_LPC_Due", "sum"),
                "LPC Collected": ("LPC_Collected", "sum"),
                "LPC pending": ("Pending_LPC", "sum"),
                "No. of loans": ("LOAN_ID", "count"),
            }
        )

    for rmk in remark_order:
        t = block(rmk).rename(columns={
            "LPC Due": f"{rmk}__LPC Due",
            "LPC Collected": f"{rmk}__LPC Collected",
            "LPC pending": f"{rmk}__LPC pending",
            "No. of loans": f"{rmk}__No. of loans",
        })
        summary_df = summary_df.merge(t, on=["ZONE_NAME", "CLUSTER_NAME"], how="left")

    summary_df = summary_df.fillna(0)

    summary_df["Total LPC Due"] = 0
    summary_df["Total LPC Collected"] = 0
    summary_df["Total LPC pending"] = 0
    summary_df["Total No.of loans"] = 0

    for rmk in remark_order:
        summary_df["Total LPC Due"] += summary_df[f"{rmk}__LPC Due"]
        summary_df["Total LPC Collected"] += summary_df[f"{rmk}__LPC Collected"]
        summary_df["Total LPC pending"] += summary_df[f"{rmk}__LPC pending"]
        summary_df["Total No.of loans"] += summary_df[f"{rmk}__No. of loans"]

    summary_df = summary_df.sort_values(["ZONE_NAME", "CLUSTER_NAME"]).reset_index(drop=True)

    # ---------- Write detail sheet by pandas (fast, low memory) ----------
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Member wise data", index=False)

    # ---------- Format summary with openpyxl ----------
    wb = load_workbook(out_path)
    if "Summary" in wb.sheetnames:
        wb.remove(wb["Summary"])
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    # Styles (local)
    thin = Side(style="thin", color="000000")
    thick = Side(style="medium", color="000000")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    font_title = Font(bold=True, size=12)
    font_head = Font(bold=True, size=10)
    font_bold = Font(bold=True)

    fill_title = PatternFill("solid", fgColor="D9EEF7")
    fill_group1 = PatternFill("solid", fgColor="FCE4D6")
    fill_group2 = PatternFill("solid", fgColor="C6EFCE")
    fill_group3 = PatternFill("solid", fgColor="EAD1F5")
    fill_group4 = PatternFill("solid", fgColor="CFEFFF")
    fill_grand = PatternFill("solid", fgColor="D9EEF7")

    def apply_border_range(ws_, r1, c1, r2, c2):
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                ws_.cell(rr, cc).border = border_thin

    def fill_range(ws_, r1, c1, r2, c2, fill):
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                ws_.cell(rr, cc).fill = fill

    def set_thick_block_outline(ws_, r1, c1, r2, c2):
        for cc in range(c1, c2 + 1):
            ws_.cell(r1, cc).border = Border(
                left=ws_.cell(r1, cc).border.left,
                right=ws_.cell(r1, cc).border.right,
                top=thick,
                bottom=ws_.cell(r1, cc).border.bottom,
            )
            ws_.cell(r2, cc).border = Border(
                left=ws_.cell(r2, cc).border.left,
                right=ws_.cell(r2, cc).border.right,
                top=ws_.cell(r2, cc).border.top,
                bottom=thick,
            )
        for rr in range(r1, r2 + 1):
            ws_.cell(rr, c1).border = Border(
                left=thick,
                right=ws_.cell(rr, c1).border.right,
                top=ws_.cell(rr, c1).border.top,
                bottom=ws_.cell(rr, c1).border.bottom,
            )
            ws_.cell(rr, c2).border = Border(
                left=ws_.cell(rr, c2).border.left,
                right=thick,
                top=ws_.cell(rr, c2).border.top,
                bottom=ws_.cell(rr, c2).border.bottom,
            )

    def paint_blocks(row_idx):
        fill_range(ws, row_idx, 2, row_idx, 5, fill_group1)
        fill_range(ws, row_idx, 6, row_idx, 9, fill_group2)
        fill_range(ws, row_idx, 10, row_idx, 13, fill_group3)
        fill_range(ws, row_idx, 14, row_idx, 17, fill_group4)

    def format_row_numbers(row_idx):
        for col in range(2, 22):
            cell = ws.cell(row_idx, col)
            cell.alignment = align_right
            cell.number_format = "0" if col in [5, 9, 13, 17, 21] else "#,##0"

    # Title row
    title = "Summary of Late Payment Charges Due, Collected & Pending as on January 21, 2026"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=21)
    c = ws.cell(1, 1, title)
    c.font = font_title
    c.alignment = align_center
    fill_range(ws, 1, 1, 1, 21, fill_title)
    apply_border_range(ws, 1, 1, 1, 21)

    group_row, sub_row, data_start = 2, 3, 4

    ws.merge_cells(start_row=group_row, start_column=1, end_row=sub_row, end_column=1)
    ws.cell(group_row, 1, "HUB/STATE").font = font_head
    ws.cell(group_row, 1).alignment = align_center
    fill_range(ws, group_row, 1, sub_row, 1, fill_title)

    group_defs = [
        ("Arrear & LPC Pending", 2, 5, fill_group1),
        ("LPC Fully Collected", 6, 9, fill_group2),
        ("No Arrear & Outstanding but LPC Pending", 10, 13, fill_group3),
        ("No Arrear but LPC Pending", 14, 17, fill_group4),
    ]
    for text, c1, c2, fill in group_defs:
        ws.merge_cells(start_row=group_row, start_column=c1, end_row=group_row, end_column=c2)
        h = ws.cell(group_row, c1, text)
        h.font = font_head
        h.alignment = align_center
        fill_range(ws, group_row, c1, group_row, c2, fill)

    subs = ["LPC Due", "LPC Collected", "LPC pending", "No.of loans"]
    for start_col, fill in [(2, fill_group1), (6, fill_group2), (10, fill_group3), (14, fill_group4)]:
        for j, sh in enumerate(subs):
            cc = ws.cell(sub_row, start_col + j, sh)
            cc.font = font_head
            cc.alignment = align_center
            cc.fill = fill

    tot_heads = ["Total LPC Due", "Total LPC Collected", "Total LPC pending", "Total No.of loans"]
    for j, th in enumerate(tot_heads):
        cell = ws.cell(sub_row, 18 + j, th)
        cell.font = font_head
        cell.alignment = align_center
        cell.fill = fill_title
    for col in range(18, 22):
        ws.cell(group_row, col).fill = fill_title

    apply_border_range(ws, group_row, 1, sub_row, 21)

    # Data with zone headers, totals, grand total
    r = data_start
    current_zone = None

    def put_block(row, start_col, rmk):
        ws.cell(r, start_col, float(row[f"{rmk}__LPC Due"]))
        ws.cell(r, start_col + 1, float(row[f"{rmk}__LPC Collected"]))
        ws.cell(r, start_col + 2, float(row[f"{rmk}__LPC pending"]))
        ws.cell(r, start_col + 3, float(row[f"{rmk}__No. of loans"]))

    def write_zone_total(zone_name, zone_df, row_idx):
        ws.cell(row_idx, 1, f"{zone_name} Total").font = font_bold
        ws.cell(row_idx, 1).alignment = align_left

        mapping = {
            2: "Arrear & LPC Pending__LPC Due",
            3: "Arrear & LPC Pending__LPC Collected",
            4: "Arrear & LPC Pending__LPC pending",
            5: "Arrear & LPC Pending__No. of loans",
            6: "LPC Fully Collected__LPC Due",
            7: "LPC Fully Collected__LPC Collected",
            8: "LPC Fully Collected__LPC pending",
            9: "LPC Fully Collected__No. of loans",
            10: "No Arrear & Outstanding but LPC Pending__LPC Due",
            11: "No Arrear & Outstanding but LPC Pending__LPC Collected",
            12: "No Arrear & Outstanding but LPC Pending__LPC pending",
            13: "No Arrear & Outstanding but LPC Pending__No. of loans",
            14: "No Arrear but LPC Pending__LPC Due",
            15: "No Arrear but LPC Pending__LPC Collected",
            16: "No Arrear but LPC Pending__LPC pending",
            17: "No Arrear but LPC Pending__No. of loans",
            18: "Total LPC Due",
            19: "Total LPC Collected",
            20: "Total LPC pending",
            21: "Total No.of loans",
        }
        for col_idx, field in mapping.items():
            v = zone_df[field].sum()
            cell = ws.cell(row_idx, col_idx, float(v))
            cell.font = font_bold
            cell.alignment = align_right
            cell.number_format = "0" if col_idx in [5, 9, 13, 17, 21] else "#,##0"

        paint_blocks(row_idx)
        apply_border_range(ws, row_idx, 1, row_idx, 21)

    for _, row in summary_df.iterrows():
        zone, cluster = row["ZONE_NAME"], row["CLUSTER_NAME"]

        if current_zone is None or zone != current_zone:
            if current_zone is not None:
                prev_zone_df = summary_df[summary_df["ZONE_NAME"] == current_zone]
                write_zone_total(current_zone, prev_zone_df, r)
                r += 1

            # Zone header row (colored blocks)
            ws.cell(r, 1, zone).font = font_bold
            ws.cell(r, 1).alignment = align_left
            paint_blocks(r)
            apply_border_range(ws, r, 1, r, 21)
            r += 1
            current_zone = zone

        ws.cell(r, 1, cluster).alignment = align_left

        put_block(row, 2, "Arrear & LPC Pending")
        put_block(row, 6, "LPC Fully Collected")
        put_block(row, 10, "No Arrear & Outstanding but LPC Pending")
        put_block(row, 14, "No Arrear but LPC Pending")

        ws.cell(r, 18, float(row["Total LPC Due"]))
        ws.cell(r, 19, float(row["Total LPC Collected"]))
        ws.cell(r, 20, float(row["Total LPC pending"]))
        ws.cell(r, 21, float(row["Total No.of loans"]))

        paint_blocks(r)
        format_row_numbers(r)
        apply_border_range(ws, r, 1, r, 21)
        r += 1

    if current_zone is not None:
        last_zone_df = summary_df[summary_df["ZONE_NAME"] == current_zone]
        write_zone_total(current_zone, last_zone_df, r)
        r += 1

    # Grand Total
    ws.cell(r, 1, "Grand Total").font = font_bold
    ws.cell(r, 1).alignment = align_left

    col_map = {
        2: "Arrear & LPC Pending__LPC Due",
        3: "Arrear & LPC Pending__LPC Collected",
        4: "Arrear & LPC Pending__LPC pending",
        5: "Arrear & LPC Pending__No. of loans",
        6: "LPC Fully Collected__LPC Due",
        7: "LPC Fully Collected__LPC Collected",
        8: "LPC Fully Collected__LPC pending",
        9: "LPC Fully Collected__No. of loans",
        10: "No Arrear & Outstanding but LPC Pending__LPC Due",
        11: "No Arrear & Outstanding but LPC Pending__LPC Collected",
        12: "No Arrear & Outstanding but LPC Pending__LPC pending",
        13: "No Arrear & Outstanding but LPC Pending__No. of loans",
        14: "No Arrear but LPC Pending__LPC Due",
        15: "No Arrear but LPC Pending__LPC Collected",
        16: "No Arrear but LPC Pending__LPC pending",
        17: "No Arrear but LPC Pending__No. of loans",
        18: "Total LPC Due",
        19: "Total LPC Collected",
        20: "Total LPC pending",
        21: "Total No.of loans",
    }
    for col_idx, field in col_map.items():
        v = summary_df[field].sum()
        cell = ws.cell(r, col_idx, float(v))
        cell.font = font_bold
        cell.alignment = align_right
        cell.number_format = "0" if col_idx in [5, 9, 13, 17, 21] else "#,##0"

    fill_range(ws, r, 1, r, 21, fill_grand)
    apply_border_range(ws, r, 1, r, 21)
    r_end = r

    # Thick outlines
    set_thick_block_outline(ws, group_row, 1, r_end, 21)
    set_thick_block_outline(ws, group_row, 2, r_end, 5)
    set_thick_block_outline(ws, group_row, 6, r_end, 9)
    set_thick_block_outline(ws, group_row, 10, r_end, 13)
    set_thick_block_outline(ws, group_row, 14, r_end, 17)
    set_thick_block_outline(ws, group_row, 18, r_end, 21)

    # Freeze / widths
    ws.freeze_panes = "B4"
    ws.column_dimensions["A"].width = 26

    # Simple width set (limited scan for speed)
    for col in range(1, 22):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 12, 12)

    wb.save(out_path)
    return df, summary_df


# ============================================================
# TAB 1: FULL AUTOMATION
# ============================================================
if selected_report == "1) Collection Efficiency Report":
    st.subheader("🏦 Collection Efficiency")
    st.caption("Upload 4 files → Run → Download DVC.xlsx (with Arrear_Advance sheet) + Hub-wise ZIP")

    c1, c2 = st.columns(2)
    with c1:
        ce_file = st.file_uploader(
            "1) Upload Collection Efficiency.xlsx",
            type=["xlsx"],
            help="Must contain a column named 'Hub' in at least one sheet.",
            key="ce"
        )
        jlg_file = st.file_uploader("2) Upload DVC JLG.xlsx", type=["xlsx"], key="jlg")
    with c2:
        il_file = st.file_uploader("3) Upload DVC IL.xlsx", type=["xlsx"], key="il")
        arrear_file = st.file_uploader(
            "4) Upload Arrear JLG & IL (ZIP/CSV/XLSX)",
            type=["zip", "csv", "xlsx"],
            help="Best: upload ZIP containing a single CSV with column 'zone_name'.",
            key="arrear"
        )

    use_csv_mode = st.checkbox("⚡ Use CSV mode for Arrear (faster)", value=True, key="csvmode")
    st.caption("Tip: If Arrear CSV is >200MB, ZIP it and upload the .zip (usually <200MB).")

    all_uploaded = all([ce_file, jlg_file, il_file, arrear_file])
    run_btn = st.button("🚀 Run Automation", disabled=not all_uploaded, use_container_width=True, key="run_full")

    if run_btn:
        try:
            progress = st.progress(0)
            status = st.empty()
            timer_box = st.empty()
            t0 = time.time()

            with tempfile.TemporaryDirectory() as workdir:
                ce_path = os.path.join(workdir, "Collection Efficiency.xlsx")
                jlg_path = os.path.join(workdir, "DVC JLG.xlsx")
                il_path = os.path.join(workdir, "DVC IL.xlsx")
                arrear_ext = os.path.splitext(arrear_file.name)[1].lower()
                arrear_path = os.path.join(workdir, f"Arrear{arrear_ext}")

                status.write("Saving uploaded files...")
                save_uploaded_file(ce_file, ce_path)
                save_uploaded_file(jlg_file, jlg_path)
                save_uploaded_file(il_file, il_path)
                save_uploaded_file(arrear_file, arrear_path)

                progress.progress(10)
                timer_box.write(f"Elapsed: {time.time()-t0:.1f}s")

                status.write("Creating DVC consolidated file...")
                consolidated_df = build_dvc_consolidate(jlg_path, il_path)

                status.write("Creating Arrear_Advance sheet (cust_id pivot)...")
                arrear_adv_df = build_arrear_advance_sheet(consolidated_df)

                dvc_output_path = os.path.join(workdir, "DVC.xlsx")
                with pd.ExcelWriter(dvc_output_path, engine="openpyxl") as writer:
                    consolidated_df.to_excel(writer, sheet_name="DVC_consolidate", index=False)
                    arrear_adv_df.to_excel(writer, sheet_name="Arrear_Advance", index=False)

                progress.progress(35)
                timer_box.write(f"Elapsed after DVC write: {time.time()-t0:.1f}s")

                status.write("Reading Collection Efficiency...")
                ce_sheets = read_ce_sheets(ce_path)

                hub_list = None
                for _, d in ce_sheets.items():
                    if "Hub" in d.columns:
                        hub_list = d["Hub"].dropna().unique()
                        break
                if hub_list is None or len(hub_list) == 0:
                    raise ValueError("No Hub values found in Collection Efficiency.xlsx (column 'Hub').")

                progress.progress(45)

                status.write("Preparing Arrear data...")
                arr_df = read_arrear_any(arrear_path, arrear_file.name, use_csv_mode)
                arr_df.columns = arr_df.columns.astype(str).str.strip()

                zone_col = None
                for c in arr_df.columns:
                    if c.strip().lower() == "zone_name":
                        zone_col = c
                        break
                if zone_col is None:
                    raise ValueError("Arrear file: Column 'zone_name' not found.")

                progress.progress(55)
                timer_box.write(f"Elapsed after Arrear ready: {time.time()-t0:.1f}s")

                status.write("Creating Hub-wise folders/files...")
                hubwise_root = os.path.join(workdir, "Hub wise")
                os.makedirs(hubwise_root, exist_ok=True)

                total_hubs = len(hub_list)
                dvc_hub_col = "hub" if "hub" in consolidated_df.columns else ("Hub" if "Hub" in consolidated_df.columns else None)

                for i, hub in enumerate(hub_list, start=1):
                    hub_folder = os.path.join(hubwise_root, clean_name_for_folder(hub))
                    os.makedirs(hub_folder, exist_ok=True)

                    wb_ce = Workbook()
                    wb_ce.remove(wb_ce.active)
                    ce_written = 0

                    for sh_name, df_sh in ce_sheets.items():
                        if "Hub" not in df_sh.columns:
                            continue

                        filtered = df_sh[df_sh["Hub"] == hub].copy()
                        if filtered.empty:
                            continue

                        filtered = recompute_grand_total(filtered, hub_col="Hub")

                        ws_ = wb_ce.create_sheet(title=safe_sheet_title(sh_name))
                        fmt_map = build_format_map(filtered, hub_col="Hub")
                        write_df_fast(ws_, filtered, fmt_map)
                        ce_written += 1

                    if ce_written == 0:
                        ws_ = wb_ce.create_sheet(title="Info")
                        ws_["A1"] = f"No Collection Efficiency data for HUB = {hub}"
                        ws_["A1"].font = Font(bold=True)

                    wb_ce.save(os.path.join(hub_folder, "Collection Efficiency.xlsx"))

                    wb_dvc = Workbook()
                    wb_dvc.remove(wb_dvc.active)

                    if dvc_hub_col is None:
                        ws_info = wb_dvc.create_sheet("Info")
                        ws_info["A1"] = "DVC output does not have hub column."
                        ws_info["A1"].font = Font(bold=True)
                    else:
                        dvc_hub_df = consolidated_df[
                            consolidated_df[dvc_hub_col].astype(str).str.strip() == str(hub).strip()
                        ].copy()

                        ws1 = wb_dvc.create_sheet("DVC_consolidate")
                        if dvc_hub_df.empty:
                            ws1["A1"] = f"No DVC data for HUB = {hub}"
                            ws1["A1"].font = Font(bold=True)
                        else:
                            fmt_map = build_format_map(dvc_hub_df, hub_col=dvc_hub_col)
                            write_df_fast(ws1, dvc_hub_df, fmt_map)

                        ws2 = wb_dvc.create_sheet("Arrear_Advance")
                        if dvc_hub_df.empty:
                            ws2["A1"] = "No data"
                            ws2["A1"].font = Font(bold=True)
                        else:
                            hub_adv = build_arrear_advance_sheet(dvc_hub_df)
                            for row in dataframe_to_rows(hub_adv, index=False, header=True):
                                ws2.append(row)
                            for cell in ws2[1]:
                                cell.font = Font(bold=True)
                            ws2.freeze_panes = "A2"
                            ws2.auto_filter.ref = ws2.dimensions

                    wb_dvc.save(os.path.join(hub_folder, "DVC.xlsx"))

                    ar_hub_df = arr_df[arr_df[zone_col].astype(str).str.strip() == str(hub).strip()].copy()
                    arrear_out_csv = os.path.join(hub_folder, "Arrear JLG & IL.csv")

                    if ar_hub_df.empty:
                        pd.DataFrame(columns=arr_df.columns).to_csv(arrear_out_csv, index=False, encoding="utf-8-sig")
                    else:
                        ar_hub_df.to_csv(arrear_out_csv, index=False, encoding="utf-8-sig")

                    progress.progress(55 + int(40 * (i / max(total_hubs, 1))))

                status.write("Packaging ZIP...")
                zip_path = os.path.join(workdir, "Hub_Wise_Output.zip")
                zip_folder(hubwise_root, zip_path)
                progress.progress(98)

                with open(dvc_output_path, "rb") as f:
                    dvc_bytes = f.read()
                with open(zip_path, "rb") as f:
                    zip_bytes = f.read()

            progress.progress(100)
            status.empty()
            timer_box.empty()
            st.success("✅ Done. Download outputs below.")

            d1, d2 = st.columns(2)
            with d1:
                st.download_button(
                    "⬇️ Download DVC.xlsx (with Arrear_Advance sheet)",
                    data=dvc_bytes,
                    file_name="DVC.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            with d2:
                st.download_button(
                    "⬇️ Download Hub-wise Output (ZIP)",
                    data=zip_bytes,
                    file_name="Hub_Wise_Output.zip",
                    mime="application/zip",
                    use_container_width=True
                )

        except Exception as e:
            st.error("Error occurred. Full details below:")
            st.exception(e)


# ============================================================
# TAB 2: ONLY DVC CONSOLIDATE + ARREAR ADVANCE REPORT
# ============================================================
if selected_report == "2) Arrear Advance Report":
    st.subheader("🏦 Arrear Advance Report")
    st.caption("Upload only JLG + IL → Download DVC.xlsx (DVC_consolidate + Arrear_Advance)")

    c1, c2 = st.columns(2)
    with c1:
        jlg2 = st.file_uploader("Upload DVC JLG.xlsx", type=["xlsx"], key="jlg2")
    with c2:
        il2 = st.file_uploader("Upload DVC IL.xlsx", type=["xlsx"], key="il2")

    run2 = st.button("🚀 Generate Report", disabled=not (jlg2 and il2), use_container_width=True, key="run2")

    if run2:
        try:
            progress = st.progress(0)
            status = st.empty()
            t0 = time.time()

            with tempfile.TemporaryDirectory() as workdir:
                jlg_path = os.path.join(workdir, "DVC JLG.xlsx")
                il_path = os.path.join(workdir, "DVC IL.xlsx")

                status.write("Saving uploaded files...")
                save_uploaded_file(jlg2, jlg_path)
                save_uploaded_file(il2, il_path)
                progress.progress(20)

                status.write("Building DVC consolidate...")
                consolidated_df = build_dvc_consolidate(jlg_path, il_path)
                progress.progress(60)

                status.write("Building Arrear_Advance (cust_id pivot)...")
                arrear_adv_df = build_arrear_advance_sheet(consolidated_df)
                progress.progress(85)

                out_path = os.path.join(workdir, "DVC.xlsx")
                with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                    consolidated_df.to_excel(writer, sheet_name="DVC_consolidate", index=False)
                    arrear_adv_df.to_excel(writer, sheet_name="Arrear_Advance", index=False)

                with open(out_path, "rb") as f:
                    out_bytes = f.read()

            progress.progress(100)
            status.empty()
            st.success(f"✅ Report generated in {time.time()-t0:.1f}s")

            st.download_button(
                "⬇️ Download DVC.xlsx (with Arrear_Advance sheet)",
                data=out_bytes,
                file_name="DVC.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        except Exception as e:
            st.error("Error occurred. Full details below:")
            st.exception(e)


# ============================================================
# TAB 3: LPC REPORT (NEW MODULE)
# ============================================================
if selected_report == "3) LPC Report":

    st.subheader("🏦LPC Report")
    st.caption("Upload LPC raw Excel → Download output with 'Summary' (formatted) + 'Member wise data' (filtered)")

    lpc_file = st.file_uploader("Upload LPC raw report (Excel)", type=["xlsx"], key="lpc_raw")
    run_lpc = st.button("🚀 Generate LPC Report", disabled=not bool(lpc_file), use_container_width=True, key="run_lpc")

    if run_lpc:
        try:
            progress = st.progress(0)
            status = st.empty()
            t0 = time.time()

            with tempfile.TemporaryDirectory() as workdir:
                in_path = os.path.join(workdir, "LPC_Raw.xlsx")
                save_uploaded_file(lpc_file, in_path)
                progress.progress(15)

                status.write("Running LPC filters + Summary formatting...")
                out_path = os.path.join(workdir, "LPC_filtered.xlsx")

                df_f, _ = lpc_generate_report(in_path, out_path, sheet_name=0)
                progress.progress(85)

                with open(out_path, "rb") as f:
                    out_bytes = f.read()

            progress.progress(100)
            status.empty()
            st.success(f"✅ LPC report generated. Filtered rows: {len(df_f)} | Time: {time.time()-t0:.1f}s")

            st.download_button(
                "⬇️ Download LPC_filtered.xlsx",
                data=out_bytes,
                file_name="LPC_filtered.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        except Exception as e:
            st.error("Error occurred. Full details below:")
            st.exception(e)
if selected_report == "4) Vault Cash Report":

    st.subheader("🏦 Vault Cash Report")

    uploaded = st.file_uploader("Upload Vault Cash Report.xlsx", type=["xlsx"], key="vault_cash")

    if uploaded:
        import tempfile, os

        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, uploaded.name)

        with open(temp_path, "wb") as f:
            f.write(uploaded.getbuffer())

        if st.button("Run Vault Cash Automation", key="run_vault"):
            try:
                run_vault_cash_report(temp_path)
                st.success("✅ Vault Cash Report generated successfully")

                with open(temp_path, "rb") as f:
                    st.download_button(
                        "⬇️ Download Updated Vault Cash Report",
                        data=f,
                        file_name="Vault Cash Report - Updated.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            except Exception as e:
                st.error("Error occurred. Full details below:")
                st.exception(e)
if selected_report == "5) Pending Collection Entry":

    # Pending Collection Entry UI only
    st.subheader("🏦 Pending Collection Entry")

    uploaded = st.file_uploader(
        "Upload Pending Collection Entry.xlsx",
        type=["xlsx"],
        key="pending_collection"
    )
    if uploaded:
        import tempfile, os

        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, uploaded.name)

        with open(temp_path, "wb") as f:
            f.write(uploaded.getbuffer())

        if st.button("Run Pending Collection Automation", key="run_pending_collection"):
            try:
                run_pending_collection_entry(temp_path)
                st.success("✅ Summary sheet created successfully")

                with open(temp_path, "rb") as f:
                    st.download_button(
                        "⬇️ Download Updated File",
                        data=f,
                        file_name="Pending Collection Entry - Updated.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            except Exception as e:
                st.error("Error occurred. Full details below:")
                st.exception(e)
if selected_report == "6) Demand Verification Report":

    st.subheader("🏦 Demand Verification Report")

    uploaded = st.file_uploader(
        "Upload Demand Verification.xlsx",
        type=["xlsx"],
        key="demand_verification"
    )

    if uploaded:
        import tempfile, os

        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, uploaded.name)

        with open(temp_path, "wb") as f:
            f.write(uploaded.getbuffer())

        if st.button("Run Demand Verification", key="run_demand_verification"):
            try:
                run_demand_verification(temp_path)
                st.success("✅ Demand Verification generated successfully")

                with open(temp_path, "rb") as f:
                    st.download_button(
                        "⬇️ Download Updated Demand Verification",
                        data=f,
                        file_name="Demand Verification - Updated.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            except Exception as e:
                st.error("❌ Error occurred. Full details below:")
                st.exception(e)
if selected_report == "7) MAP Ledger Difference":

    st.subheader("📌 MAP vs Ledger – Difference Report")

    map_file = st.file_uploader("Upload MAP.xlsx (must have Consolidated sheet)", type=["xlsx"], key="map_xlsx")
    ledger_file = st.file_uploader("Upload MAP Ledger.xlsx", type=["xlsx"], key="map_ledger_xlsx")

    if map_file and ledger_file:
        import tempfile

        temp_dir = tempfile.mkdtemp()

        map_path = os.path.join(temp_dir, map_file.name)
        ledger_path = os.path.join(temp_dir, ledger_file.name)

        with open(map_path, "wb") as f:
            f.write(map_file.getbuffer())

        with open(ledger_path, "wb") as f:
            f.write(ledger_file.getbuffer())

        if st.button("Run MAP Ledger Difference", key="run_map_diff"):
            try:
                run_map_ledger_difference(map_path, ledger_path)
                st.success("✅ Difference sheet generated in MAP Ledger file!")

                with open(ledger_path, "rb") as f:
                    st.download_button(
                        "⬇️ Download Updated MAP Ledger.xlsx",
                        data=f,
                        file_name="MAP Ledger - Updated.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            except Exception as e:
                st.error("❌ Error occurred. Details below:")
                st.exception(e)
if selected_report == "8) Excess Amount":
    st.subheader("💰 Excess Amount Received")

    il = st.file_uploader("Upload Repayment Summary IL.xlsx", type=["xlsx"], key="ex_il")
    jlg = st.file_uploader("Upload Repayment Summary JLG.xlsx", type=["xlsx"], key="ex_jlg")
    esc = st.file_uploader("Upload Escalation.xlsx", type=["xlsx"], key="ex_esc")

    if il and jlg and esc:
        import tempfile, os

        temp_dir = tempfile.mkdtemp()

        il_path = os.path.join(temp_dir, "Repayment Summary IL.xlsx")
        jlg_path = os.path.join(temp_dir, "Repayment Summary JLG.xlsx")
        esc_path = os.path.join(temp_dir, "Escalation.xlsx")

        with open(il_path, "wb") as f:
            f.write(il.getbuffer())
        with open(jlg_path, "wb") as f:
            f.write(jlg.getbuffer())
        with open(esc_path, "wb") as f:
            f.write(esc.getbuffer())

if st.button("Run Excess Amount Automation"):
    try:
        out_file = run_excess_amount_from_streamlit(il_path, jlg_path, esc_path, temp_dir)
        st.success("Generated successfully")

        with open(out_file, "rb") as f:
            st.download_button(
                "⬇️ Download Consolidated Excess Amount",
                data=f,
                file_name=os.path.basename(out_file),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    except Exception as e:
        st.error("Error occurred. Full details below:")
        st.exception(e)

if selected_report == "9) CPP Payable":

    st.subheader("📑 CPP Payable vs CPP Ledger")

    cpp_payable = st.file_uploader(
        "Upload CPP Payable.xlsx",
        type=["xlsx"],
        key="cpp_payable"
    )

    cpp_ledger = st.file_uploader(
        "Upload CPP Ledger.xlsx",
        type=["xlsx"],
        key="cpp_ledger"
    )

    if cpp_payable and cpp_ledger:
        import tempfile, os

        temp_dir = tempfile.mkdtemp()
        cpp_payable_path = os.path.join(temp_dir, cpp_payable.name)
        cpp_ledger_path = os.path.join(temp_dir, cpp_ledger.name)

        with open(cpp_payable_path, "wb") as f:
            f.write(cpp_payable.getbuffer())

        with open(cpp_ledger_path, "wb") as f:
            f.write(cpp_ledger.getbuffer())

        if st.button("Run CPP Difference Report", key="run_cpp_diff"):
            try:
                run_cpp_payable_vs_cpp_ledger_difference(
                    cpp_payable_path,
                    cpp_ledger_path
                )

                st.success("✅ CPP Difference report generated successfully")

                with open(cpp_ledger_path, "rb") as f:
                    st.download_button(
                        "⬇️ Download Updated CPP Ledger.xlsx",
                        data=f,
                        file_name="CPP Ledger - Updated.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

            except Exception as e:
                st.error("Error occurred while processing")
                st.exception(e)
if selected_report == "10) CMS and UPI Recon":

    st.subheader("🏦 CMS Recon")

    cms_format = st.file_uploader("Upload CMS Recon Sheet.xlsx", type=["xlsx"], key="cms_format")
    cms_ledger = st.file_uploader("Upload CMS Ledger.xlsx", type=["xlsx"], key="cms_ledger")
    cms_zip = st.file_uploader("Upload BS Folder (ZIP of all statements)", type=["zip"], key="cms_zip")

    if cms_format and cms_ledger and cms_zip:
        import tempfile, os, zipfile

        temp_dir = tempfile.mkdtemp()

        # Save uploaded files
        format_path = os.path.join(temp_dir, "CMS_Recon_Sheet.xlsx")
        ledger_path = os.path.join(temp_dir, "CMS_Ledger.xlsx")
        zip_path = os.path.join(temp_dir, "BS.zip")

        with open(format_path, "wb") as f:
            f.write(cms_format.getbuffer())

        with open(ledger_path, "wb") as f:
            f.write(cms_ledger.getbuffer())

        with open(zip_path, "wb") as f:
            f.write(cms_zip.getbuffer())

        # Extract ZIP -> statements folder
        statements_folder = os.path.join(temp_dir, "BS")
        os.makedirs(statements_folder, exist_ok=True)

        with zipfile.ZipFile(zip_path, "r") as z:
            z.extractall(statements_folder)

        # Handle case where ZIP contains an inner "BS" folder
        inner_bs = os.path.join(statements_folder, "BS")
        if os.path.isdir(inner_bs):
            statements_folder = inner_bs

        output_path = os.path.join(temp_dir, "Consolidate CMS Recon.xlsx")

        if st.button("Run CMS Recon", key="run_cms_recon"):
            try:
                run_cms_recon_streamlit(format_path, statements_folder, ledger_path, output_path)

                st.success("✅ CMS Recon created successfully")

                with open(output_path, "rb") as f:
                    st.download_button(
                        "⬇️ Download Consolidate CMS Recon.xlsx",
                        data=f,
                        file_name="Consolidate CMS Recon.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

            except Exception as e:
                st.error("❌ Error occurred. Details below:")
                st.exception(e)


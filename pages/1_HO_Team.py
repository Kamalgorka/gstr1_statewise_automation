import streamlit as st
import os
import zipfile
import pandas as pd
from openpyxl import load_workbook, Workbook
import warnings
import re
import shutil
import time
from datetime import datetime
from openpyxl.styles import Font, Alignment
from ui import load_global_css
from ho_daybook_core import run_daybook_from_uploaded_files
load_global_css()

warnings.filterwarnings("ignore", category=UserWarning)
st.set_page_config(
    page_title="GSTR-1 State-wise Automation",
    page_icon="üìä",
    layout="wide"
)

# ‚úÖ ADD BELOW THIS LINE
ho_report = st.selectbox(
    "",
    [
        "1) GSTR-1 State-wise Automation",
        "2) HO DayBook Automation"
    ],
    key="ho_report_select"
)
st.markdown("<br>", unsafe_allow_html=True)


BASE_DIR = os.path.dirname(os.path.abspath(__file__))

UPLOAD_FOLDER = os.path.join("/tmp", "uploads")
OUTPUT_ROOT = os.path.join("/tmp", "outputs")
TEMPLATE_FILE = os.path.join(BASE_DIR, "..", "GSTR1_Excel_Workbook_Template.xlsx")
LOG_FILE = os.path.join(OUTPUT_ROOT, "audit_log.xlsx")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_ROOT, exist_ok=True)

if not os.path.exists(TEMPLATE_FILE):
    st.error(f"‚ùå Template file not found in app folder: {TEMPLATE_FILE}")
    st.stop()


def log_activity(action, month, filename):
    now = datetime.now()

    if not os.path.exists(LOG_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Log"
        ws.append(["Date", "Time", "Action", "Month", "File Name", "Machine Name"])
        wb.save(LOG_FILE)

    wb = load_workbook(LOG_FILE)
    ws = wb.active

    ws.append([
        now.strftime("%d-%m-%Y"),
        now.strftime("%H:%M:%S"),
        action,
        month,
        filename,
        os.environ.get("COMPUTERNAME", "Unknown")
    ])

    wb.save(LOG_FILE)


STATE_LIST = [
    "PUNJAB", "RAJASTHAN", "HARYANA", "CHANDIGARH", "BIHAR",
    "UTTAR PRADESH", "HIMACHAL PRADESH", "JHARKHAND", "GUJARAT",
    "UTTARAKHAND", "MADHYA PRADESH", "WEST BENGAL", "ODISHA",
    "JAMMU & KASHMIR", "ANDHRA PRADESH", "CHHATTISGARH", "ASSAM", "MAHARASHTRA"
]

STATE_CODE_MAP = {
    "JAMMU & KASHMIR": "01-Jammu & Kashmir",
    "HIMACHAL PRADESH": "02-Himachal Pradesh",
    "PUNJAB": "03-Punjab",
    "CHANDIGARH": "04-Chandigarh",
    "UTTARAKHAND": "05-Uttarakhand",
    "HARYANA": "06-Haryana",
    "DELHI": "07-Delhi",
    "RAJASTHAN": "08-Rajasthan",
    "UTTAR PRADESH": "09-Uttar Pradesh",
    "BIHAR": "10-Bihar",
    "SIKKIM": "11-Sikkim",
    "ARUNACHAL PRADESH": "12-Arunachal Pradesh",
    "NAGALAND": "13-Nagaland",
    "MANIPUR": "14-Manipur",
    "MIZORAM": "15-Mizoram",
    "TRIPURA": "16-Tripura",
    "MEGHALAYA": "17-Meghalaya",
    "ASSAM": "18-Assam",
    "WEST BENGAL": "19-West Bengal",
    "JHARKHAND": "20-Jharkhand",
    "ODISHA": "21-Odisha",
    "CHHATTISGARH": "22-Chhattisgarh",
    "MADHYA PRADESH": "23-Madhya Pradesh",
    "GUJARAT": "24-Gujarat",
    "DAMAN & DIU": "25-Daman & Diu",
    "DADRA & NAGAR HAVELI & DAMAN & DIU": "26-Dadra & Nagar Haveli & Daman & Diu",
    "MAHARASHTRA": "27-Maharashtra",
    "KARNATAKA": "29-Karnataka",
    "GOA": "30-Goa",
    "LAKSHADWEEP": "31-Lakshadweep",
    "KERALA": "32-Kerala",
    "TAMIL NADU": "33-Tamil Nadu",
    "PUDUCHERRY": "34-Puducherry",
    "ANDAMAN & NICOBAR ISLANDS": "35-Andaman & Nicobar Islands",
    "TELANGANA": "36-Telangana",
    "ANDHRA PRADESH": "37-Andhra Pradesh",
    "LADAKH": "38-Ladakh",
    "FOREIGN COUNTRY": "96-Foreign Country",
    "OTHER TERRITORY": "97-Other Territory"
}


def format_place_of_supply(state_val):
    if pd.isna(state_val):
        return ""
    key = str(state_val).strip().upper()
    return STATE_CODE_MAP.get(key, state_val)


def clean_hsn(x):
    x = str(x).strip()
    if x.endswith(".0"):
        x = x[:-2]
    return x


def filter_state_month(df, state_upper, month_str):
    if "State" not in df.columns:
        return df.iloc[0:0]
    s = df["State"].astype(str).str.upper().str.strip()
    mask = (s == state_upper)
    if "Month" in df.columns:
        m = df["Month"].astype(str).str.lower().str.strip()
        mask = mask & m.str.contains(month_str.lower(), na=False)
    return df[mask]


def extract_last_num(x):
    nums = re.findall(r'\d+', str(x))
    return int(nums[-1]) if nums else 0


# ======================================================
# ‚úÖ 1) FILE VALIDATION (Sheets + Key Columns)
# ======================================================
REQUIRED_SHEETS = [
    "GSTR1", "B2B", "CD Note Reg", "DN Note Reg", "Exempt Income",
    "B2C PF", "CD Note Unreg", "B2C Onboarding"
]

REQUIRED_COLUMNS = {
    "B2B": ["State", "Invoice Type", "Invoice Number", "Invoice Date", "Total Transaction Value", "State Place of Supply"],
    "CD Note Reg": ["State", "Month"],
    "DN Note Reg": ["State", "Month"],
    "B2C PF": ["State", "Month", "LPF"],
    "CD Note Unreg": ["State", "Month"],
    "Exempt Income": ["Row Labels", "Sum of Collection Intrest", "Month"],
    # "GSTR1" and "B2C Onboarding" kept open because formats vary
}


def validate_excel(file_path: str):
    errors = []
    warnings_list = []

    try:
        xl = pd.ExcelFile(file_path)
        sheet_names = xl.sheet_names
    except Exception as e:
        return [f"‚ùå Unable to read Excel file. Error: {e}"], []

    missing_sheets = [s for s in REQUIRED_SHEETS if s not in sheet_names]
    if missing_sheets:
        errors.append("‚ùå Missing required sheet(s): " + ", ".join(missing_sheets))

    # If mandatory sheets missing, stop further checks
    if errors:
        return errors, warnings_list

    # Column checks (lightweight: read only headers)
    for sh, cols in REQUIRED_COLUMNS.items():
        try:
            df_head = pd.read_excel(file_path, sheet_name=sh, nrows=1)
            missing_cols = [c for c in cols if c not in df_head.columns]
            if missing_cols:
                errors.append(f"‚ùå Sheet '{sh}' missing column(s): {', '.join(missing_cols)}")
        except Exception as e:
            errors.append(f"‚ùå Could not read sheet '{sh}'. Error: {e}")

    # Extra practical warnings
    # Check State values exist in B2B
    try:
        df_states = pd.read_excel(file_path, sheet_name="B2B", usecols=["State"])
        if df_states["State"].dropna().empty:
            warnings_list.append("‚ö†Ô∏è 'B2B' sheet has empty 'State' column. State-wise output may be blank.")
    except Exception:
        pass

    return errors, warnings_list


# ======================================================
# ‚úÖ 2) CACHING FOR READING ALL SHEETS
# ======================================================
@st.cache_data(show_spinner=False)
def read_all_sheets_cached(file_bytes: bytes):
    # Use BytesIO so cache is based on file content
    from io import BytesIO
    bio = BytesIO(file_bytes)

    df_gstr1 = pd.read_excel(bio, sheet_name="GSTR1")
    bio.seek(0)
    df_b2b = pd.read_excel(bio, sheet_name="B2B")
    bio.seek(0)
    df_cd = pd.read_excel(bio, sheet_name="CD Note Reg")
    bio.seek(0)
    df_dn = pd.read_excel(bio, sheet_name="DN Note Reg")
    bio.seek(0)
    df_exempt = pd.read_excel(bio, sheet_name="Exempt Income")
    bio.seek(0)
    df_b2c_pf = pd.read_excel(bio, sheet_name="B2C PF")
    bio.seek(0)
    df_cdunreg = pd.read_excel(bio, sheet_name="CD Note Unreg")
    bio.seek(0)
    df_b2c_onboard = pd.read_excel(bio, sheet_name="B2C Onboarding")

    # normalize once
    if "Month" in df_b2c_pf.columns:
        df_b2c_pf["Month_norm"] = df_b2c_pf["Month"].astype(str).str.lower().str.strip()
    else:
        df_b2c_pf["Month_norm"] = ""

    return df_gstr1, df_b2b, df_cd, df_dn, df_exempt, df_b2c_pf, df_cdunreg, df_b2c_onboard


# ======================================================
# ‚úÖ 3) PROGRESS BAR + ETA (per state)
# ======================================================
def run_gstr_process(
    df_gstr1, df_b2b, df_cd, df_dn, df_exempt,
    df_b2c_pf, df_cdunreg, df_b2c_onboard,
    MONTH, MONTH_NORM,
    TEMPLATE_FILE, OUTPUT_FOLDER,
    progress_bar=None,
    status_box=None,
    eta_box=None
):
    total_states = len(STATE_LIST)
    start_all = time.time()

    for i, state in enumerate(STATE_LIST, start=1):
        state_start = time.time()

        wb = load_workbook(TEMPLATE_FILE)

        for sheet_name in wb.sheetnames:
            if sheet_name.lower() not in ["exemp", "docs"]:
                ws_tmp = wb[sheet_name]
                ws_tmp.delete_rows(4)

        df_state_b2b = filter_state_month(df_b2b, state, MONTH)
        df_cd_state = filter_state_month(df_cd, state, MONTH)
        df_dn_state = filter_state_month(df_dn, state, MONTH)

        # B2B
        ws = wb["b2b"]
        r = 4
        for _, row in df_state_b2b.iterrows():
            if str(row["Invoice Type"]).upper() == "B2B":
                ws[f"A{r}"] = row.get("Customer Billing GSTIN", "")
                ws[f"B{r}"] = row.get("Customer Billing Name", "")
                ws[f"C{r}"] = row.get("Invoice Number", "")
                ws[f"D{r}"] = row.get("Invoice Date", "")
                ws[f"E{r}"] = row.get("Total Transaction Value", 0)
                ws[f"F{r}"] = format_place_of_supply(row.get("State Place of Supply", ""))
                ws[f"G{r}"] = "N"
                ws[f"H{r}"] = ""
                ws[f"I{r}"] = "Regular B2B"
                ws[f"J{r}"] = ""
                ws[f"K{r}"] = 18
                ws[f"L{r}"] = row.get("Item Taxable Value", row.get("Total Transaction Value", 0))
                ws[f"M{r}"] = ""
                r += 1

        # B2CL
        ws = wb["b2cl"]
        r = 4
        for _, row in df_state_b2b.iterrows():
            if str(row["Invoice Type"]).upper() == "B2CL":
                ws[f"A{r}"] = row.get("Invoice Number", "")
                ws[f"B{r}"] = row.get("Invoice Date", "")
                ws[f"C{r}"] = row.get("Total Transaction Value", 0)
                ws[f"D{r}"] = row.get("State Place of Supply", "")
                ws[f"E{r}"] = ""
                ws[f"F{r}"] = 18
                ws[f"G{r}"] = row.get("Item Taxable Value", row.get("Total Transaction Value", 0))
                ws[f"H{r}"] = ""
                ws[f"I{r}"] = ""
                r += 1

        # B2CS
        ws = wb["b2cs"]
        r = 4

        df_b2cs_1 = df_state_b2b[df_state_b2b["Invoice Type"].astype(str).str.upper() == "B2CS"].copy()
        if not df_b2cs_1.empty:
            df_b2cs_1["POS"] = df_b2cs_1["State Place of Supply"].astype(str).str.upper()
            df_b2cs_1["TAX"] = df_b2cs_1.get("Item Taxable Value", df_b2cs_1["Total Transaction Value"])
            df_b2cs_1 = df_b2cs_1[["POS", "TAX"]]
        else:
            df_b2cs_1 = pd.DataFrame(columns=["POS", "TAX"])

        df_b2c_pf_state = df_b2c_pf[
            (df_b2c_pf["State"].astype(str).str.upper() == state) &
            (df_b2c_pf["Month_norm"].str.contains(MONTH_NORM, na=False))
        ].copy()

        df_b2cs_2 = pd.DataFrame(columns=["POS", "TAX"])
        if not df_b2c_pf_state.empty:
            df_b2cs_2["POS"] = df_b2c_pf_state["State"].astype(str).str.upper()
            df_b2cs_2["TAX"] = df_b2c_pf_state["LPF"].fillna(0)

        df_b2cs_all = pd.concat([df_b2cs_1, df_b2cs_2], ignore_index=True)
        df_b2cs_grp = df_b2cs_all.groupby("POS", as_index=False)["TAX"].sum() if not df_b2cs_all.empty else pd.DataFrame(columns=["POS", "TAX"])

        df_cdun_b2cs = df_cdunreg[
            (df_cdunreg["State"].astype(str).str.upper() == state) &
            (df_cdunreg["Month"].astype(str).str.lower().str.contains(MONTH_NORM, na=False))
        ].copy()

        df_cdun_b2cs_adj = pd.DataFrame(columns=["POS", "CD_TAX"])
        if not df_cdun_b2cs.empty:
            df_cdun_b2cs["POS"] = df_cdun_b2cs["State"].astype(str).str.upper()
            df_cdun_b2cs_adj = (
                df_cdun_b2cs
                .groupby("POS", as_index=False)["Item Taxable Value"]
                .sum()
                .rename(columns={"Item Taxable Value": "CD_TAX"})
            )

        df_b2cs_final = df_b2cs_grp.merge(df_cdun_b2cs_adj, on="POS", how="left").fillna(0) if not df_b2cs_grp.empty else pd.DataFrame(columns=["POS", "TAX", "CD_TAX"])
        if not df_b2cs_final.empty:
            df_b2cs_final["TAX"] = (df_b2cs_final["TAX"] - df_b2cs_final["CD_TAX"]).clip(lower=0)

            for _, row in df_b2cs_final.iterrows():
                ws[f"A{r}"] = "OE"
                ws[f"B{r}"] = format_place_of_supply(row["POS"])
                ws[f"C{r}"] = ""
                ws[f"D{r}"] = 18
                ws[f"E{r}"] = row["TAX"]
                ws[f"F{r}"] = ""
                ws[f"G{r}"] = ""
                ws[f"H{r}"] = ""
                r += 1

        # CDNR
        ws = wb["cdnr"]
        r = 4
        notes_df = pd.concat([df_cd_state, df_dn_state], ignore_index=True)
        for _, row in notes_df.iterrows():
            note_type_full = str(row.get("Credit(C)/Debit(D) Note Type", "")).strip().upper()
            if note_type_full.startswith("C"):
                mapped_type = "C"
            elif note_type_full.startswith("D"):
                mapped_type = "D"
            else:
                mapped_type = ""

            gstin = row.get("Customer Billing GSTIN", row.get("My GSTIN", ""))
            receiver_name = row.get("Customer Billing Name", "")
            note_number = row.get("Credit/Debit Note Number", "")
            note_date = row.get("Credit/Debit Note Date", "")
            pos = row.get("State Place of Supply", row.get("State", state))
            total_val = row.get("Total Transaction Value", row.get("Invoice Amount", 0))
            taxable_val = row.get("Item Taxable Value", total_val)

            ws[f"A{r}"] = gstin
            ws[f"B{r}"] = receiver_name
            ws[f"C{r}"] = note_number
            ws[f"D{r}"] = note_date
            ws[f"E{r}"] = mapped_type
            ws[f"F{r}"] = format_place_of_supply(pos)
            ws[f"G{r}"] = "N"
            ws[f"H{r}"] = "Regular"
            ws[f"I{r}"] = total_val
            ws[f"J{r}"] = ""
            ws[f"K{r}"] = 18
            ws[f"L{r}"] = taxable_val
            ws[f"M{r}"] = ""
            r += 1

        # EXEMPT
        ws = wb["exemp"]
        exempt_val = None
        if "Row Labels" in df_exempt.columns and "Sum of Collection Intrest" in df_exempt.columns:
            mask_state = df_exempt["Row Labels"].astype(str).str.upper().str.strip() == state
            if "Month" in df_exempt.columns:
                m = df_exempt["Month"].astype(str).str.lower().str.strip()
                mask_month = m.str.contains(MONTH_NORM, na=False)
                mask = mask_state & mask_month
            else:
                mask = mask_state

            if mask.any():
                exempt_val = df_exempt.loc[mask, "Sum of Collection Intrest"].sum()

        for col in ["B", "C", "D"]:
            for row_idx in range(4, 8):
                ws[f"{col}{row_idx}"] = ""

        ws["C7"] = float(exempt_val) if (exempt_val and not pd.isna(exempt_val) and exempt_val != 0) else ""

        # HSN (b2b)
        ws = wb["hsn (b2b)"]
        r = 4
        df_state_hsn = df_state_b2b[df_state_b2b["Invoice Type"].astype(str).str.upper() == "B2B"][[
            "HSN or SAC Code", "Total Transaction Value", "Item Taxable Value",
            "IGST Amount", "CGST Amount", "SGST Amount"
        ]].copy() if not df_state_b2b.empty else pd.DataFrame(columns=[
            "HSN or SAC Code", "Total Transaction Value", "Item Taxable Value",
            "IGST Amount", "CGST Amount", "SGST Amount"
        ])

        if not df_state_hsn.empty:
            df_state_hsn["HSN"] = df_state_hsn["HSN or SAC Code"].astype(str).apply(clean_hsn)

            df_hsn_b2b_grp = df_state_hsn.groupby("HSN", as_index=False)[[
                "Total Transaction Value", "Item Taxable Value",
                "IGST Amount", "CGST Amount", "SGST Amount"
            ]].sum()
        else:
            df_hsn_b2b_grp = pd.DataFrame(columns=[
                "HSN", "Total Transaction Value", "Item Taxable Value",
                "IGST Amount", "CGST Amount", "SGST Amount"
            ])

        df_cdreg_state = df_cd[
            (df_cd["State"].astype(str).str.upper() == state) &
            (df_cd["Month"].astype(str).str.lower().str.contains(MONTH_NORM, na=False))
        ].copy()

        if not df_cdreg_state.empty and "HSN or SAC Code" in df_cdreg_state.columns:
            df_cdreg_state["HSN"] = df_cdreg_state["HSN or SAC Code"].apply(clean_hsn)
            df_cdreg_state = df_cdreg_state.rename(columns={
                "Total Transaction Value": "cd_total",
                "Item Taxable Value": "cd_taxable",
                "IGST Amount": "cd_igst",
                "CGST Amount": "cd_cgst",
                "SGST Amount": "cd_sgst"
            })
            cd_group = df_cdreg_state.groupby("HSN", as_index=False)[[
                "cd_total", "cd_taxable", "cd_igst", "cd_cgst", "cd_sgst"
            ]].sum()
        else:
            cd_group = pd.DataFrame(columns=["HSN", "cd_total", "cd_taxable", "cd_igst", "cd_cgst", "cd_sgst"])

        df_final_hsn = df_hsn_b2b_grp.merge(cd_group, on="HSN", how="left").fillna(0) if not df_hsn_b2b_grp.empty else pd.DataFrame(columns=[
            "HSN", "Total Transaction Value", "Item Taxable Value", "IGST Amount", "CGST Amount", "SGST Amount",
            "cd_total", "cd_taxable", "cd_igst", "cd_cgst", "cd_sgst"
        ])

        if not df_final_hsn.empty:
            df_final_hsn["Total Transaction Value"] -= df_final_hsn["cd_total"]
            df_final_hsn["Item Taxable Value"] -= df_final_hsn["cd_taxable"]
            df_final_hsn["IGST Amount"] -= df_final_hsn["cd_igst"]
            df_final_hsn["CGST Amount"] -= df_final_hsn["cd_cgst"]
            df_final_hsn["SGST Amount"] -= df_final_hsn["cd_sgst"]

            df_final_hsn[[
                "Total Transaction Value", "Item Taxable Value",
                "IGST Amount", "CGST Amount", "SGST Amount"
            ]] = df_final_hsn[[
                "Total Transaction Value", "Item Taxable Value",
                "IGST Amount", "CGST Amount", "SGST Amount"
            ]].clip(lower=0)

            for _, row in df_final_hsn.iterrows():
                ws[f"A{r}"] = row["HSN"]
                ws[f"E{r}"] = row["Total Transaction Value"]
                ws[f"F{r}"] = row["Item Taxable Value"]
                ws[f"G{r}"] = 18
                ws[f"H{r}"] = row["IGST Amount"]
                ws[f"I{r}"] = row["CGST Amount"]
                ws[f"J{r}"] = row["SGST Amount"]
                r += 1

        # ======================================================
        # ‚úÖ HSN (B2C) ‚Äì FULL BLOCK (WITH EXEMPT INCOME HSN)
        # ======================================================
        # Safe sheet getter (avoids sheet-name mismatch issue)
        def get_sheet(wb, target_name):
            def norm(x):
                return str(x).strip().lower().replace(" ", "")
            t = norm(target_name)
            for s in wb.sheetnames:
                if norm(s) == t:
                    return wb[s]
            return None

        ws = get_sheet(wb, "hsn (b2c)")
        if ws is None:
            raise ValueError(f"Template is missing sheet: 'hsn (b2c)'. Available: {wb.sheetnames}")

        r = 4

        # -----------------------------------------
        # 1Ô∏è‚É£ B2C data from B2B sheet (non-B2B invoices)
        # -----------------------------------------
        needed_cols = [
            "HSN or SAC Code", "Total Transaction Value", "Item Taxable Value",
            "IGST Amount", "CGST Amount", "SGST Amount"
        ]

        df_b2c_part1 = df_state_b2b[
            df_state_b2b["Invoice Type"].astype(str).str.upper() != "B2B"
        ].copy()

        # If required columns missing, keep part1 empty (no crash)
        if all(c in df_b2c_part1.columns for c in needed_cols):
            df_b2c_part1 = df_b2c_part1[needed_cols].copy()
            df_b2c_part1.rename(columns={"HSN or SAC Code": "HSN"}, inplace=True)
            df_b2c_part1["HSN"] = df_b2c_part1["HSN"].astype(str).apply(clean_hsn)
        else:
            df_b2c_part1 = pd.DataFrame(columns=[
                "HSN", "Total Transaction Value", "Item Taxable Value",
                "IGST Amount", "CGST Amount", "SGST Amount"
            ])

        # -----------------------------------------
        # 2Ô∏è‚É£ B2C PF data
        # -----------------------------------------
        df_b2c_pf_state = df_b2c_pf[
            (df_b2c_pf["State"].astype(str).str.upper() == state) &
            (df_b2c_pf["Month_norm"].str.contains(MONTH_NORM, na=False))
        ].copy()

        df_b2c_part2 = pd.DataFrame(columns=[
            "HSN", "Total Transaction Value", "Item Taxable Value",
            "IGST Amount", "CGST Amount", "SGST Amount"
        ])

        # This part only works if your B2C PF sheet has HSN column
        if (not df_b2c_pf_state.empty) and ("HSN" in df_b2c_pf_state.columns):
            df_b2c_part2 = pd.DataFrame({
                "HSN": df_b2c_pf_state["HSN"].astype(str).apply(clean_hsn),
                "Total Transaction Value": (
                    df_b2c_pf_state.get("LPF", 0).fillna(0) +
                    df_b2c_pf_state.get("IGST", 0).fillna(0) +
                    df_b2c_pf_state.get("CGST", 0).fillna(0) +
                    df_b2c_pf_state.get("SGST", 0).fillna(0)
                ),
                "Item Taxable Value": df_b2c_pf_state.get("LPF", 0).fillna(0),
                "IGST Amount": df_b2c_pf_state.get("IGST", 0).fillna(0),
                "CGST Amount": df_b2c_pf_state.get("CGST", 0).fillna(0),
                "SGST Amount": df_b2c_pf_state.get("SGST", 0).fillna(0)
            })

        # -----------------------------------------
        # 3Ô∏è‚É£ Combine B2C HSN data
        # -----------------------------------------
        df_hsn_b2c_all = pd.concat([df_b2c_part1, df_b2c_part2], ignore_index=True)

        if df_hsn_b2c_all.empty:
            df_hsn_b2c_grp = pd.DataFrame(columns=[
                "HSN", "Total Transaction Value", "Item Taxable Value",
                "IGST Amount", "CGST Amount", "SGST Amount"
            ])
        else:
            df_hsn_b2c_grp = df_hsn_b2c_all.groupby("HSN", as_index=False)[[
                "Total Transaction Value", "Item Taxable Value",
                "IGST Amount", "CGST Amount", "SGST Amount"
            ]].sum()

        # -----------------------------------------
        # 4Ô∏è‚É£ CD NOTE UNREG adjustment (POS independent, HSN wise)
        # -----------------------------------------
        df_cdun_state = filter_state_month(df_cdunreg, state, MONTH).copy()

        df_cdun_hsn = pd.DataFrame(columns=[
            "HSN", "Total Transaction Value", "Item Taxable Value",
            "IGST Amount", "CGST Amount", "SGST Amount"
        ])

        if (not df_cdun_state.empty) and all(c in df_cdun_state.columns for c in needed_cols):
            df_cdun_hsn = df_cdun_state[needed_cols].copy()
            df_cdun_hsn.rename(columns={"HSN or SAC Code": "HSN"}, inplace=True)
            df_cdun_hsn["HSN"] = df_cdun_hsn["HSN"].astype(str).apply(clean_hsn)

        if df_cdun_hsn.empty:
            df_cdun_grp = pd.DataFrame(columns=df_hsn_b2c_grp.columns)
        else:
            df_cdun_grp = df_cdun_hsn.groupby("HSN", as_index=False)[[
                "Total Transaction Value", "Item Taxable Value",
                "IGST Amount", "CGST Amount", "SGST Amount"
            ]].sum()

        if df_hsn_b2c_grp.empty:
            df_final_b2c_hsn = df_hsn_b2c_grp.copy()
        else:
            df_final_b2c_hsn = df_hsn_b2c_grp.merge(
                df_cdun_grp,
                on="HSN",
                how="left",
                suffixes=("", "_CD")
            ).fillna(0)

            df_final_b2c_hsn["Total Transaction Value"] -= df_final_b2c_hsn.get("Total Transaction Value_CD", 0)
            df_final_b2c_hsn["Item Taxable Value"] -= df_final_b2c_hsn.get("Item Taxable Value_CD", 0)
            df_final_b2c_hsn["IGST Amount"] -= df_final_b2c_hsn.get("IGST Amount_CD", 0)
            df_final_b2c_hsn["CGST Amount"] -= df_final_b2c_hsn.get("CGST Amount_CD", 0)
            df_final_b2c_hsn["SGST Amount"] -= df_final_b2c_hsn.get("SGST Amount_CD", 0)

            df_final_b2c_hsn = df_final_b2c_hsn[[
                "HSN", "Total Transaction Value", "Item Taxable Value",
                "IGST Amount", "CGST Amount", "SGST Amount"
            ]].copy()

        df_final_b2c_hsn[[
            "Total Transaction Value", "Item Taxable Value",
            "IGST Amount", "CGST Amount", "SGST Amount"
        ]] = df_final_b2c_hsn[[
            "Total Transaction Value", "Item Taxable Value",
            "IGST Amount", "CGST Amount", "SGST Amount"
        ]].clip(lower=0)

        # -----------------------------------------
        # 5Ô∏è‚É£ ADD EXEMPT INCOME HSN (NEW REQUIREMENT)
        # -----------------------------------------
        df_exempt_state = df_exempt[
            df_exempt["Row Labels"].astype(str).str.upper().str.strip() == state
        ].copy()

        if "Month" in df_exempt_state.columns:
            df_exempt_state = df_exempt_state[
                df_exempt_state["Month"].astype(str).str.lower().str.contains(MONTH_NORM, na=False)
            ]

        if (not df_exempt_state.empty) and ("HSN or SAC Code" in df_exempt_state.columns):
            df_exempt_hsn = pd.DataFrame({
                "HSN": df_exempt_state["HSN or SAC Code"].astype(str).apply(clean_hsn),
                "Total Transaction Value": df_exempt_state["Sum of Collection Intrest"].fillna(0),
                "Item Taxable Value": df_exempt_state["Sum of Collection Intrest"].fillna(0),
                "IGST Amount": 0,
                "CGST Amount": 0,
                "SGST Amount": 0
            })

            df_final_b2c_hsn = pd.concat([df_final_b2c_hsn, df_exempt_hsn], ignore_index=True)
            df_final_b2c_hsn = df_final_b2c_hsn.groupby("HSN", as_index=False).sum()

        # -----------------------------------------
        # 6Ô∏è‚É£ Write to Excel
        # -----------------------------------------
        for _, row in df_final_b2c_hsn.iterrows():
            ws[f"A{r}"] = row["HSN"]
            ws[f"E{r}"] = row["Total Transaction Value"]
            ws[f"F{r}"] = row["Item Taxable Value"]

            # Rate = 0 only for HSN 997114, else 18
            ws[f"G{r}"] = 0 if str(row["HSN"]).strip() == "997114" else 18

            ws[f"H{r}"] = row["IGST Amount"]
            ws[f"I{r}"] = row["CGST Amount"]
            ws[f"J{r}"] = row["SGST Amount"]
            r += 1


        # DOCS (as per your existing block)
        ws = wb["docs"]

        df_docs_b2b = df_state_b2b[df_state_b2b.get("Invoice Number").notna()].copy() if not df_state_b2b.empty else pd.DataFrame()
        if not df_docs_b2b.empty and "Invoice Number" in df_docs_b2b.columns:
            df_docs_b2b["InvNum"] = df_docs_b2b["Invoice Number"].apply(extract_last_num)
            min_inv = df_docs_b2b["InvNum"].min()
            max_inv = df_docs_b2b["InvNum"].max()
            total_count = max_inv - min_inv + 1
            sr_from = df_docs_b2b[df_docs_b2b["InvNum"] == min_inv]["Invoice Number"].iloc[0]
            sr_to = df_docs_b2b[df_docs_b2b["InvNum"] == max_inv]["Invoice Number"].iloc[0]
        else:
            sr_from = sr_to = ""
            total_count = 0

        ws["A4"] = "Invoices for outward supply"
        ws["B4"] = sr_from
        ws["C4"] = sr_to
        ws["D4"] = total_count
        ws["E4"] = 0

        df_b2c_pf_state_prev = df_b2c_pf[
            (df_b2c_pf["State"].astype(str).str.upper() == state) &
            (~df_b2c_pf["Month_norm"].str.contains(MONTH_NORM, na=False))
        ]
        df_b2c_pf_state_curr = df_b2c_pf[
            (df_b2c_pf["State"].astype(str).str.upper() == state) &
            (df_b2c_pf["Month_norm"].str.contains(MONTH_NORM, na=False))
        ]

        T_prev = len(df_b2c_pf_state_prev)
        T_curr = len(df_b2c_pf_state_curr)
        sr_from_b2c = T_prev + 1 if T_curr > 0 else ""
        sr_to_b2c = T_prev + T_curr if T_curr > 0 else ""

        ws["A5"] = "Invoices for outward supply"
        ws["B5"] = sr_from_b2c
        ws["C5"] = sr_to_b2c
        ws["D5"] = T_curr
        ws["E5"] = 0

        ws["A6"] = "Invoices for outward supply"
        ws["B6"] = ""
        ws["C6"] = ""
        ws["D6"] = ""
        ws["E6"] = 0

        df_cd_unreg_prev = df_cdunreg[
            (df_cdunreg["State"].astype(str).str.upper() == state) &
            (~df_cdunreg["Month"].astype(str).str.lower().str.contains(MONTH_NORM, na=False))
        ]
        df_cd_unreg_curr = df_cdunreg[
            (df_cdunreg["State"].astype(str).str.upper() == state) &
            (df_cdunreg["Month"].astype(str).str.lower().str.contains(MONTH_NORM, na=False))
        ]
        CD_prev = len(df_cd_unreg_prev)
        CD_curr = len(df_cd_unreg_curr)
        sr_from_cd = CD_prev + 1 if CD_curr > 0 else ""
        sr_to_cd = CD_prev + CD_curr if CD_curr > 0 else ""

        ws["A7"] = "Credit Note"
        ws["B7"] = sr_from_cd
        ws["C7"] = sr_to_cd
        ws["D7"] = CD_curr
        ws["E7"] = 0

        df_dn_reg_prev = df_dn[
            (df_dn["State"].astype(str).str.upper() == state) &
            (~df_dn["Month"].astype(str).str.lower().str.contains(MONTH_NORM, na=False))
        ]
        df_dn_reg_curr = df_dn[
            (df_dn["State"].astype(str).str.upper() == state) &
            (df_dn["Month"].astype(str).str.lower().str.contains(MONTH_NORM, na=False))
        ]
        DN_prev = len(df_dn_reg_prev)
        DN_curr = len(df_dn_reg_curr)
        sr_from_dn = DN_prev + 1 if DN_curr > 0 else ""
        sr_to_dn = DN_prev + DN_curr if DN_curr > 0 else ""

        ws["A8"] = "Debit Note"
        ws["B8"] = sr_from_dn
        ws["C8"] = sr_to_dn
        ws["D8"] = DN_curr
        ws["E8"] = 0

        # Keep your docs formatting fix

        ws_docs = wb["docs"]
        ws_docs["H4"] = ""
        for col in range(1, ws_docs.max_column + 1):
            cell = ws_docs.cell(row=5, column=col)
            cell.font = Font(bold=False)
            cell.alignment = Alignment(
                vertical="center",
                horizontal=cell.alignment.horizontal,
                wrap_text=False
            )
        ws_docs.row_dimensions[5].height = None

        out_path = os.path.join(OUTPUT_FOLDER, f"GSTR1_{MONTH}_{state}.xlsx")
        os.makedirs(OUTPUT_FOLDER, exist_ok=True)
        wb.save(out_path)

        # --- Progress updates ---
        if status_box is not None:
            status_box.info(f"‚úÖ Completed: {state}  ({i}/{total_states})")

        if progress_bar is not None:
            progress_bar.progress(i / total_states)

        # ETA based on average time per state so far
        if eta_box is not None:
            elapsed = time.time() - start_all
            avg_per_state = elapsed / max(i, 1)
            remaining = avg_per_state * (total_states - i)
            eta_box.caption(f"‚è≥ Estimated remaining processing time: ~{int(remaining)} sec")

    return


# ======================================================
# UI
# ======================================================
st.markdown("""
<style>
.big-title { font-size: 30px; font-weight: 700; color: #0A3D62; }
.sub-title { font-size: 16px; color: #576574; }
.card { background-color: #F8F9FA; padding: 20px; border-radius: 12px;
        box-shadow: 0px 2px 6px rgba(0,0,0,0.08); }
</style>
""", unsafe_allow_html=True)
# ======================================================
# UI
# ======================================================
st.markdown("""
<style>
.big-title { font-size: 30px; font-weight: 700; color: #0A3D62; }
.sub-title { font-size: 16px; color: #576574; }
.card { background-color: #F8F9FA; padding: 20px; border-radius: 12px;
        box-shadow: 0px 2px 6px rgba(0,0,0,0.08); }
</style>
""", unsafe_allow_html=True)

if ho_report == "1) GSTR-1 State-wise Automation":

    st.markdown('<div class="big-title">üìë GSTR-1 State-wise Automation</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-title">Upload ‚Üí Validate ‚Üí Process ‚Üí Download</div>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    col1, col2 = st.columns([2, 1])

    with col1:
        uploaded_file = st.file_uploader("üìÇ Upload GSTR1 format.xlsx", type=["xlsx"])

    with col2:
        month = st.selectbox("üìÖ Select Month", [
            "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December"
        ])

    st.markdown('</div>', unsafe_allow_html=True)

    # placeholders for progress UI
    progress_bar = st.progress(0)
    status_box = st.empty()
    eta_box = st.empty()

    if st.button("üöÄ Generate State-wise Excel Files", use_container_width=True):
        if uploaded_file is None:
            st.error("‚ùå Please upload GSTR1 format.xlsx")
            st.stop()

        # Save upload to disk (keeping your existing approach)
        gstr1_path = os.path.join(UPLOAD_FOLDER, "GSTR1_format.xlsx")
        file_bytes = uploaded_file.getvalue()
        with open(gstr1_path, "wb") as f:
            f.write(file_bytes)

        # ‚úÖ Validation step (with clear error detection)
        errs, warns = validate_excel(gstr1_path)
        if errs:
            st.error("File validation failed. Please fix below issue(s):")
            for e in errs:
                st.write(e)
            st.stop()

        if warns:
            st.warning("Validation warning(s):")
            for w in warns:
                st.write(w)

        MONTH = month
        MONTH_NORM = month.lower()

        OUTPUT_FOLDER = os.path.join(OUTPUT_ROOT, MONTH)
        if os.path.exists(OUTPUT_FOLDER):
            shutil.rmtree(OUTPUT_FOLDER)
        os.makedirs(OUTPUT_FOLDER)

        # Reset progress UI
        progress_bar.progress(0)
        status_box.info("‚è≥ Starting processing...")
        eta_box.caption("")

        # ‚úÖ Cached read of sheets
        with st.spinner("üì• Reading Excel (cached for faster repeats)..."):
            df_gstr1, df_b2b, df_cd, df_dn, df_exempt, df_b2c_pf, df_cdunreg, df_b2c_onboard = read_all_sheets_cached(file_bytes)

        # ‚úÖ Progress bar + ETA while processing
        with st.spinner("‚öôÔ∏è Processing state-wise files..."):
            run_gstr_process(
                df_gstr1, df_b2b, df_cd, df_dn, df_exempt,
                df_b2c_pf, df_cdunreg, df_b2c_onboard,
                MONTH, MONTH_NORM,
                TEMPLATE_FILE, OUTPUT_FOLDER,
                progress_bar=progress_bar,
                status_box=status_box,
                eta_box=eta_box
            )

        status_box.success("üéâ Processing completed!")

        # üéâ Center Popup "It's Done" (auto-disappears)
        done_popup = st.empty()
        done_popup.markdown(
            """
            <div id="done-popup" style="
                position: fixed;
                top: 50%;
                left: 50%;
                transform: translate(-50%, -50%);
                background: linear-gradient(90deg, #00b894, #0984e3);
                color: white;
                padding: 28px 36px;
                border-radius: 16px;
                font-size: 24px;
                font-weight: 700;
                text-align: center;
                box-shadow: 0px 12px 30px rgba(0,0,0,0.35);
                z-index: 9999;
                animation: fadein 0.6s;
            ">
                üéâ It‚Äôs Done!<br>
                <div style="font-size: 15px; font-weight: 400; margin-top: 8px;">
                    All State-wise Files Generated Successfully.<br>
                    You can now download your files.
                </div>
            </div>

            <style>
            @keyframes fadein {
                from { opacity: 0; transform: translate(-50%, -60%); }
                to   { opacity: 1; transform: translate(-50%, -50%); }
            }
            </style>
            """,
            unsafe_allow_html=True
        )
        time.sleep(4)
        done_popup.empty()

        # ‚úÖ ZIP creation (show ‚Äúpreparing download‚Äù)
        zip_path = os.path.join(OUTPUT_ROOT, f"GSTR1_{MONTH}.zip")

        with st.spinner("üì¶ Preparing ZIP for download..."):
            t0 = time.time()
            with zipfile.ZipFile(zip_path, "w") as zipf:
                for file in os.listdir(OUTPUT_FOLDER):
                    zipf.write(os.path.join(OUTPUT_FOLDER, file), arcname=file)

            log_activity(action="GSTR-1 Generated", month=MONTH, filename=f"GSTR1_{MONTH}.zip")
            t1 = time.time()

        zip_size_mb = os.path.getsize(zip_path) / (1024 * 1024)
        st.success("‚úÖ ZIP Ready for download.")
        st.caption(f"üì¶ ZIP Size: {zip_size_mb:.2f} MB | ZIP creation time: {t1 - t0:.1f} sec")

        est_seconds_10mbps = zip_size_mb / 1.25 if zip_size_mb > 0 else 0
        st.caption(f"‚è¨ Estimated download time (example): ~{int(est_seconds_10mbps)} sec at ~10 Mbps. (Actual depends on internet speed)")

        st.markdown("## üì• Download Summary")

        for file in os.listdir(OUTPUT_FOLDER):
            if not file.endswith(".xlsx"):
                continue

            state_name = file.replace(f"GSTR1_{MONTH}_", "").replace(".xlsx", "")
            file_path = os.path.join(OUTPUT_FOLDER, file)

            c1, c2 = st.columns([3, 1])
            with c1:
                st.markdown(f"‚úÖ **{state_name}**")
            with c2:
                with open(file_path, "rb") as f:
                    if st.download_button("‚¨á Download", data=f, file_name=file):
                        log_activity(action="State File Downloaded", month=MONTH, filename=file)

        with open(zip_path, "rb") as z:
            st.download_button("‚¨á Download GSTR1 ZIP", z, file_name=f"GSTR1_{MONTH}.zip")

elif ho_report == "2) HO DayBook Automation":

    st.markdown('<div class="big-title">üè¶ HO DayBook Automation</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-title">Upload ‚Üí Process ‚Üí Download</div>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    col1, col2 = st.columns([2, 2])

    with col1:
        coa_file = st.file_uploader("üìÇ Upload COA.xlsx", type=["xlsx"], key="ho_db_coa")

    with col2:
        zip_file = st.file_uploader("üìÇ Upload Statement.zip", type=["zip"], key="ho_db_zip")

    st.markdown('</div>', unsafe_allow_html=True)

    if st.button("üöÄ Generate HO DayBook", use_container_width=True):
        if coa_file is None or zip_file is None:
            st.error("‚ùå Please upload both COA.xlsx and Statement.zip")
            st.stop()

        try:
            with st.spinner("‚öôÔ∏è Processing DayBook..."):
                output_bytes = run_daybook_from_uploaded_files(
                    coa_file.getvalue(),
                    zip_file.getvalue()
                )

            st.success("‚úÖ HO DayBook Generated!")

            st.download_button(
                "‚¨á Download HO_DayBook_AllStatements.xlsx",
                data=output_bytes,
                file_name="HO_DayBook_AllStatements.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        except Exception as e:
            st.error("‚ùå Error occurred. Details below:")
            st.exception(e)
